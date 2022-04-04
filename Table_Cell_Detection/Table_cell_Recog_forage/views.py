"""
Version to be maintained
pandas : 1.3.5          --> 1.2.4 will not work
numpy  >=1.17.3         --> we are using 1.20.1
"""

from __future__ import absolute_import
from __future__ import division
from __future__ import print_function
from django.http import HttpResponse, JsonResponse
from rest_framework.views import APIView
from .libs.response_json import JResponseJSON
from .libs import statuscode
import fitz
import cv2
import pandas as pd
import numpy as np
import os
import shutil
import io
from pathlib import Path
import warnings
import logging
import pickle
import json
import PIL.Image as Image

from openpyxl.workbook import Workbook
from openpyxl import load_workbook

import pyodbc as pyodbc

# ------------- SQL Database connection -----------------
details = {
 }
connect_string = 'DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={server};PORT=1443;DATABASE={database};UID={username};PWD={password}'.format(**details)
connection = pyodbc.connect(connect_string,unicode_results = True)
# Initialise the Cursor
cursor = connection.cursor()

#--------------------------------------------------------------

warnings.filterwarnings('ignore')

'''------------------------------------ Setting Logger --------------------------------'''

logging.basicConfig(filename="LogFile_table_cell.log",
                    format='%(asctime)s %(message)s'
                    )
logger = logging.getLogger()
logger.setLevel(logging.INFO)

''' ------------------------------------ using MMCV as model --------------------------- '''
from mmdet.apis import init_detector, inference_detector

try:
    FILE = Path(__file__).resolve()
    ROOT = FILE.parents[0]  # YOLOv5 root directory
    ROOT = ROOT.replace("\\", "/")
    ROOT = ROOT +'/table_cell' + '/'
except:
    ROOT = os.getcwd()
    ROOT = ROOT.replace("\\", "/")
    ROOT = ROOT +'/table_cell' + '/'
logger.info('\n\n\t\t--------------------- Root Directory is : {} -------------------\t\n'.format(ROOT))


class PyMuPdf:

    def __init__(self):

        self.doc = None
        self.source = None
        self.result_dir = None
        self.lookup_category_id_annotation_df = pd.DataFrame.from_dict(
            {'category_id': [0, 1, 2, 3, 4, 5, 6],
             'name': ['semi_grided_table', 'grided_table', 'gridless_table', 'key_value', 'cell', 'special_cell', 'spanning_cell'],
             'supercategory': ['table', 'table', 'table', 'table', 'sentence', 'sentence', 'sentence']
             })

        self.lookup_detections_df = None
        self.column_detection = None

        self.cell_model = None
        self.cell_det_threshold = 0.8
        self.table_model = None
        self.table_det_threshold = 0.9
        self.column_model = None
        self.column_det_threshold = 0.9

        self.excel_dir = r'/home/ubuntu/Downloads/efs/4i/table_cell_service/'
        self.db_status = None

    def get_width(self, x0, x1):
        return int(x1-x0)

    def get_heigth(self, h0, h1):
        return int(h1-h0)

    def get_sourceid(self, row):
        return self.source

    def draw_detetion_save_img(self, tbl_lst, img_p, page_num, name_of_file='_result.jpeg', color=(255, 0, 255), result_save=False):
        '''
        tbl_list is a standard format used across the table detection code
        tbl_list format : [ x1,y1, x2, y2, labl, conf]

        or it could be a dataframe having all detections stored
        '''

        if isinstance(img_p, np.ndarray):
            img = img_p.copy()
        else:
            img = cv2.imread(img_p)
        if isinstance(tbl_lst, list):
            for det in tbl_lst:
                x1 = det[0]
                y1 = det[1]
                x2 = det[2]
                y2 = det[3]
                label = det[4]
                img = cv2.rectangle(img, (x1,y1), (x2,y2), color=color, thickness=2)
            if result_save:
                cv2.imwrite(self.result_dir + str(page_num)+name_of_file, img)

        if isinstance(tbl_lst, pd.DataFrame):
            for i in range(len(tbl_lst)):
                cv2.rectangle(img, (tbl_lst.loc[i, 'x'], tbl_lst.loc[i, 'y']), (tbl_lst.loc[i, 'x'] + tbl_lst.loc[i, 'w'], tbl_lst.loc[i, 'y'] + tbl_lst.loc[i, 'h']), color=color, thickness=2)
            if result_save:
                cv2.imwrite(self.result_dir + str(page_num) + name_of_file, img)

    def get_lookup_detection_frame(self, det_lst, ref_frame):
        '''
        This Function creates a dataframe of detections made by the model
        Input:
                det_lst         :   List of detections in table_list format
                ref_frame       :   self.lookup_category_id_annotation_df
        Output:
                df_             :   Dataframe of table_list Detection
        '''
        df = pd.DataFrame(columns=['id','x', 'y', 'w', 'h', 'x2', 'y2', 'label'])
        for d in det_lst:
            x1, y1, x2, y2, label, conf = d
            if conf> 0.8:
                df = df.append({'x': x1, 'y': y1, 'w': x2 - x1, 'h': y2 - y1, 'x2': x2, 'y2': y2, 'label': label, 'conf':conf}, ignore_index=True)
                df = df.sort_values(['y', 'x'], ascending=True)

        df_ = pd.merge(df, ref_frame, how='left', left_on='label', right_on='name')
        df_ = df_.drop_duplicates()
        df_.reset_index(drop=True, inplace=True)
        df_['id'] = df_.index

        # overlap information check between various cells in detetion frame
        df_2 = self.overlap_of_detections_check(df_)
        return df_2

    def get_undetected_parts_img(self, img_p, det_lst, page_num, save_img=True):
        '''
        This fucntions saves and return image in ndarray of parts not detected by the model
        Parms   :
        img         : image path
        det_list    : List of all detections by model in table_list format [[]]
        page_num    : page num

        Return  :
        final       : Masked image which contains only undetected parts
        '''
        img = cv2.imread(img_p, 0)
        mask = np.ones(shape=img.shape[:2], dtype="uint8")
        for d in det_lst:
            x1, y1, x2, y2, label, conf = d
            if label not in ['semi_grided_table', 'grided_table', 'gridless_table', 'key_value']:
                cv2.rectangle(mask, (x1, y1), (x2, y2), 255, -1)
        final = cv2.addWeighted(img, 1, mask, 1, 0)
        if save_img:
            cv2.imwrite(self.result_dir + str(page_num) + '_undetected.jpeg', final)
        return final

    def get_undetected_parts_bbox_v1_old(self, input_df, det_df, ref_frame):

        '''
        this function identifies the undetected cells and their coordinates.
        These new bbox cordinates are updated to model detetion output.
        '''

        # subscripting input dataframe for cells which did not get detected but is inside table.
        new_df = input_df[(input_df['is_in_cell'] == 0) & (input_df['is_in_table'] == 1)]  #
        # creating a copy of detection dataframe
        det_df_copy = det_df.copy()

        for i in new_df.index:
            undet_x = new_df.loc[i, 'x']
            undet_y = new_df.loc[i, 'y']
            undet_w = new_df.loc[i, 'w']
            undet_h = new_df.loc[i, 'h']
            undet_x2 = undet_x + undet_w
            undet_y2 = undet_y + undet_h
            for j in range(len(det_df)):
                det_x = det_df.loc[j, 'x']
                det_x2 = det_df.loc[j, 'x2']
                det_w = det_df.loc[j, 'w']
                label = det_df.loc[j, 'label']
                cate_id = det_df.loc[j, 'category_id']
                super_cat = det_df.loc[j, 'supercategory']

                # check if x values are in range and possible overlap is there on x-axis values
                if label not in ['semi_grided_table', 'grided_table', 'gridless_table', 'key_value']:
                    if undet_w < det_w:
                        if ((undet_x > det_x) and (undet_x2 > det_x)) and ((undet_x < det_x2) and (undet_x2 < det_x2)):
                            # changes made to input data frame directly
                            input_df.loc[i, 'is_in_cell'] = 1
                            input_df.loc[i, 'cell_id'] = j
                            input_df.loc[i, 'cell_type'] = label
                            input_df.loc[i, 'c_x'] = det_x
                            input_df.loc[i, 'c_y'] = undet_y
                            input_df.loc[i, 'c_w'] = det_w
                            input_df.loc[i, 'c_h'] = undet_h

                            det_df_copy = det_df_copy.append(
                                {'x': det_x, 'y': undet_y, 'w': det_w, 'h': undet_h, 'x2': det_x2,
                                 'y2': undet_y2, 'label': label, 'category_id': cate_id, 'name': label,
                                 'supercategory': super_cat}, ignore_index=True)

                    elif det_w < undet_w:
                        if ((det_x > undet_x) and (det_x2 > undet_x)) and ((det_x < undet_x2) and (det_x2 < undet_x2)):
                            input_df.loc[i, 'is_in_cell'] = 1
                            input_df.loc[i, 'cell_id'] = j
                            input_df.loc[i, 'cell_type'] = label
                            input_df.loc[i, 'c_x'] = det_x
                            input_df.loc[i, 'c_y'] = undet_y
                            input_df.loc[i, 'c_w'] = undet_w
                            input_df.loc[i, 'c_h'] = undet_h

                            det_df_copy = det_df_copy.append(
                                {'x': det_x, 'y': undet_y, 'w': undet_w, 'h': undet_h, 'x2': det_x + undet_w,
                                 'y2': undet_y2, 'label': label, 'category_id': cate_id, 'name': label,
                                 'supercategory': super_cat}, ignore_index=True)

        det_df_copy = det_df_copy.drop_duplicates()
        det_df_copy = det_df_copy.sort_values(['x', 'w', 'y'])
        det_df_copy.reset_index(drop=True, inplace=True)

        '''
        the det_df_copy may have some extra detections whih are not required , so a further process will be deployed to reduce it
        '''
        cell_det = input_df.loc[input_df['is_in_cell'] == 1]
        cell_det = cell_det[['c_x', 'c_y', 'c_w', 'c_h', 'cell_type']]
        cell_det = cell_det.drop_duplicates(keep='first').reset_index(drop=True)
        cell_det = cell_det.rename(columns={'c_x': 'x', 'c_y': 'y', 'c_w': 'w', 'c_h': 'h', 'cell_type': 'name'})

        table_det = input_df.loc[input_df['is_in_table'] == 1]
        table_det = table_det[['t_x', 't_y', 't_w', 't_h', 'table_type']]
        table_det = table_det.drop_duplicates(keep='first').reset_index(drop=True)
        table_det = table_det.rename(columns={'t_x': 'x', 't_y': 'y', 't_w': 'w', 't_h': 'h', 'table_type': 'name'})

        final = table_det.append(cell_det)
        final_ = pd.merge(final, ref_frame, how='left', left_on='name', right_on='name')

        return input_df, final_

    def get_undetected_parts_bbox_v2_old(self, input_df, det_df, ref_frame):
        '''
        this function identifies the undetected cells and their coordinates.
        These new bbox cordinates are updated to model detetion output.
        '''

        det_df_copy = det_df.copy()
        det_tab = det_df.loc[det_df['supercategory'] == 'table']
        det_df = det_df.loc[det_df['supercategory'] == 'sentence']
        det_df.sort_values(by=['x', 'w'], ascending=True, inplace=True)

        # removing all detections which are outside the table
        for ind in det_df_copy.index:
            y1 = det_df_copy.loc[ind, 'y']
            y2 = det_df_copy.loc[ind, 'y2']
            for inde in det_tab.index:
                table_y1 = det_tab.loc[inde, 'y']
                table_y2 = det_tab.loc[inde, 'y2']
                if y1 > table_y2 or y2 < table_y1:
                    det_df.drop(ind, axis=0, inplace=True)

        # undeteted_annotations update
        new_df = input_df[(input_df['is_in_cell'] == 0) & (input_df['is_in_table'] == 1)]
        # creating an empty dataframe
        undet_frame = pd.DataFrame(columns=['x1', 'y1', 'x2', 'y2'])
        check_ids_done = []
        # making annotation-coordinates for undetected values which may overlap with detections  of model.
        # result is saved in undet_frame above
        for i in new_df.index:
            undet_x = new_df.loc[i, 'x']
            undet_y = new_df.loc[i, 'y']
            undet_w = new_df.loc[i, 'w']
            undet_h = new_df.loc[i, 'h']
            undet_x2 = undet_x + undet_w
            undet_y2 = undet_y + undet_h
            for j in det_df.index:
                det_x = det_df.loc[j, 'x']
                det_x2 = det_df.loc[j, 'x2']
                # possiible overlap condiions with previous detections
                if (det_x - 5 < undet_x < det_x2 + 5) or (det_x - 5 < undet_x2 < det_x2 + 5) or (
                        undet_x <= det_x and undet_x2 >= det_x2) or (undet_x >= det_x and undet_x2 <= det_x2):
                    undet_frame = undet_frame.append(
                        {'x1': min(det_x - 5, undet_x - 5), 'y1': undet_y, 'x2': max(det_x2 + 5, undet_x2 + 5),
                         'y2': undet_y2}, ignore_index=True)
                    # here we update the index values for which coordinate overlap was found between
                    # undetected words and model detection
                    check_ids_done.append(i)
                    break
        undet_frame.drop_duplicates(inplace=True)
        undet_frame.sort_values(by=['y1', 'y2', 'x2'], inplace=True)
        undet_frame = undet_frame.reset_index(drop=True, inplace=False)

        #  here we identify the blocks of individual line based on y1 and y2
        # for undetected detection values so that minimum and max of box can be fixed avoiding overlaps
        unique_y1s = undet_frame['y1'].unique()
        su_blk = []
        for y in unique_y1s:
            block = []
            ctr = 0
            temp_undet_frame = undet_frame.loc[undet_frame['y1'] == y]
            temp_undet_frame = temp_undet_frame.reset_index(drop=False)
            for i in temp_undet_frame.index:
                up_right = temp_undet_frame.loc[i, 'x2']  # x0 coordinate of a
                if i + 1 != len(temp_undet_frame):
                    down_left = temp_undet_frame.loc[i + 1, 'x1']

                if i == 0:
                    block.append(ctr)
                # same column check using our understanding that if 2 words or cells are in same line but diff columns,
                # x2 of left word/cell will be less than x1 of right word/cell
                elif i + 1 != len(temp_undet_frame) and up_right < down_left:
                    block.append(ctr)
                    ctr = ctr + 1
                else:
                    block.append(ctr)
            temp_undet_frame['block'] = block
            su_blk.extend(block)

            # getting subset frames from the subset frame to update values in main undet_frame
            temp_blk_uni = temp_undet_frame['block'].unique()
            for blk in temp_blk_uni:
                temp_blk_frame = temp_undet_frame.loc[temp_undet_frame['block'] == blk]
                temp_blk_frame = temp_blk_frame.reset_index(drop=True)
                # at this stage we know we have a frame where detection belong to same block number ie column AND
                # same line number ie y1 y2 coordinates, therefore :
                for i in range(len(temp_blk_frame)):
                    index_v = temp_blk_frame.loc[i, 'index']
                    min_x1 = min(temp_blk_frame['x1'])
                    max_x2 = max(temp_blk_frame['x2'])
                    undet_frame.loc[index_v, 'x1'] = min_x1
                    undet_frame.loc[index_v, 'x2'] = max_x2
        # undet_frame['block'] = su_blk
        undet_frame.drop_duplicates(inplace=True)
        undet_frame.sort_values(by=['y1', 'y2', 'x2'], inplace=True)
        undet_frame = undet_frame.reset_index(drop=True, inplace=False)
        # undet_frame = undet_frame.drop(['block'], axis=1)

        # Leftover IDs are index values which did not have an overlap with  ML detections
        # we only consider cells where width of text in cell is greater than 10
        # to avoid unwanted values like '.'
        leftover_ids = list(set(list(new_df.index)) - set(check_ids_done))
        leftover_new_df = new_df.loc[leftover_ids]
        leftover_new_df = leftover_new_df.loc[leftover_new_df['w'] > 10]
        leftover_new_df = leftover_new_df[['x', 'y', 'w', 'h']]

        x2 = leftover_new_df['x'] + leftover_new_df['w']
        y2 = leftover_new_df['y'] + leftover_new_df['h']
        leftover_new_df['x2'] = x2
        leftover_new_df['y2'] = y2
        leftover_new_df.sort_values(by=['x', 'x2', 'y'], inplace=True)
        leftover_new_df = leftover_new_df.reset_index(drop=False)

        # Repeating the step of block creation as done above
        # here we mostly have cells or values which maybe in same column but complete column was not detected
        block = []
        ctr = 0
        for i in leftover_new_df.index:
            up_right = leftover_new_df.loc[i, 'x2']
            if i + 1 != len(leftover_new_df):
                down_left = leftover_new_df.loc[i + 1, 'x']
            if i == 0:
                block.append(ctr)
            # same column check
            elif i + 1 != len(leftover_new_df) and up_right < down_left:
                block.append(ctr)
                ctr = ctr + 1
            else:
                block.append(ctr)
        leftover_new_df['block'] = block
        temp_blk_uni = leftover_new_df['block'].unique()
        for blk in temp_blk_uni:
            temp_blk_frame = leftover_new_df.loc[leftover_new_df['block'] == blk]
            temp_blk_frame = temp_blk_frame.reset_index(drop=True)
            # at this stage we know we have a frame where detection belong to same block number ie column AND
            # same line number ie y1 y2 coordinates, therefore :
            for i in range(len(temp_blk_frame)):
                min_x1 = min(temp_blk_frame['x'])
                max_x2 = max(temp_blk_frame['x2'])
                undet_frame = undet_frame.append(
                    {'x1': min_x1, 'y1': temp_blk_frame.loc[i, 'y'], 'x2': max_x2, 'y2': temp_blk_frame.loc[i, 'y2']}, ignore_index=True)
        undet_frame.drop_duplicates(inplace=True)
        undet_frame.sort_values(by=['y1', 'y2', 'x2'], inplace=True)
        undet_frame = undet_frame.reset_index(drop=True, inplace=False)

        # reformating the UN-detections frame
        undet_frame['w'] = undet_frame['x2'] - undet_frame['x1']
        undet_frame['h'] = undet_frame['y2'] - undet_frame['y1']
        undet_frame['label'] = 'cell'
        undet_frame['category_id'] = 4
        undet_frame['name'] = 'cell'
        undet_frame['supercategory'] = 'sentence'
        undet_frame = undet_frame[['x1', 'y1', 'w', 'h', 'x2', 'y2', 'label', 'category_id', 'name', 'supercategory']]
        undet_frame = undet_frame.rename(columns={'x1': 'x', 'y1': 'y'})
        # appending the undetected annotations to detection frame
        det_df_copy = det_df_copy.append(undet_frame, ignore_index=True)
        # marking overlaps again based on new overlaps
        input_df = self.df_to_table_df_v2(input_df=input_df, det_df=det_df_copy)

        # now removing overlap of detections between coordinates detected by model and our undetected
        tab = det_df_copy.copy()
        tab = tab.loc[tab['supercategory'] == 'table']

        src = det_df_copy.copy()
        src = src.loc[src['name'] == 'cell']
        src = src.sort_values(by=['y', 'x'])
        src = src.reset_index(drop=True)

        for ind in src.index:
            com_x1 = src.loc[ind, 'x']
            com_y1 = src.loc[ind, 'y']
            com_x2 = src.loc[ind, 'x2']
            com_y2 = src.loc[ind, 'y2']
            com_w = src.loc[ind, 'w']
            com_h = src.loc[ind, 'h']
            for jind in src.index:
                if ind != jind:
                    jcom_x1 = src.loc[jind, 'x']
                    jcom_y1 = src.loc[jind, 'y']
                    jcom_x2 = src.loc[jind, 'x2']
                    jcom_y2 = src.loc[jind, 'y2']
                    jcom_w = src.loc[jind, 'w']
                    jcom_h = src.loc[jind, 'h']
                    # (R1[0] >= R2[2]) or (R1[2] <= R2[0]) or (R1[3] <= R2[1]) or (R1[1] >= R2[3])
                    if self.isRectangleOverlap([com_x1, com_y1, com_x2, com_y2], [jcom_x1, jcom_y1, jcom_x2, jcom_y2]):
                        x1 = min(src.loc[ind, 'x'], src.loc[jind, 'x'])
                        y1 = min(src.loc[ind, 'y'], src.loc[jind, 'y'])
                        x2 = max(src.loc[ind, 'x2'], src.loc[jind, 'x2'])
                        y2 = max(src.loc[ind, 'y2'], src.loc[jind, 'y2'])
                        w = max(src.loc[ind, 'w'], src.loc[jind, 'w'])
                        h = max(src.loc[ind, 'h'], src.loc[jind, 'h'])

                        src.loc[ind, 'x'] = x1
                        src.loc[jind, 'x'] = x1
                        src.loc[ind, 'x2'] = x2
                        src.loc[jind, 'x2'] = x2

                        src.loc[ind, 'y2'] = y2
                        src.loc[jind, 'y2'] = y2
                        src.loc[ind, 'y'] = y1
                        src.loc[jind, 'y'] = y1

                        src.loc[ind, 'w'] = w
                        src.loc[jind, 'w'] = w
                        src.loc[ind, 'h'] = h
                        src.loc[jind, 'h'] = h

        src.drop_duplicates(inplace=True)
        src = src.reset_index(drop=True)
        det_df_copy = src.append(tab)
        det_df_copy.reset_index(drop=True, inplace=True)

        return input_df, det_df_copy

    def get_undetected_parts_bbox(self, input_df, det_df, col_det):

        det_df_copy = det_df.copy()

        det_tab = det_df.loc[det_df['supercategory'] == 'table']
        det_sen = det_df.loc[det_df['supercategory'] == 'sentence']

        # removing all detections outside table
        # if sentence detections overlap with table coordinates we keep it else Removed
        for sind in det_sen.index:
            to_drop = True
            R1 = [det_sen.loc[sind, 'x'], det_sen.loc[sind, 'y'], det_sen.loc[sind, 'x2'], det_sen.loc[sind, 'y2']]
            for tind in det_tab.index:
                R2 = [det_tab.loc[tind, 'x'], det_tab.loc[tind, 'y'], det_tab.loc[tind, 'x2'], det_tab.loc[tind, 'y2']]
                boo, _, _ = self.isRectangleOverlap(R1, R2)
                if boo is True:
                    # if overlap is true we do not want to drop this index , hence
                    to_drop = False
            if to_drop is True:
                det_df_copy.drop(sind, inplace=True)

        # det_df_copy  --->  Cells detected by Model inside Table Frame
        det_df_copy.drop_duplicates(inplace=True)
        det_df_copy.sort_values(by=['x'], inplace=True)
        det_df_copy.reset_index(drop=True, inplace=True)

        # undetected_text identified using Metadata
        new_df = input_df[(input_df['is_in_cell'] == 0) & (input_df['is_in_table'] == 1)]

        # creating an empty dataframe
        undet_frame = pd.DataFrame(columns=['id','x', 'y', 'w', 'h', 'x2', 'y2', 'label', 'conf', 'category_id', 'name', 'supercategory', 'overlap'])

        # ---------------------- now we find overlap with exsisting Detections -------------

        # Extract metadata infor of words that did not get detected and find X-overlaps
        ctr = len(det_df) + 2
        for i in new_df.index:
            undet_x = new_df.loc[i, 'x']
            undet_w = new_df.loc[i, 'w']
            undet_x2 = undet_x + undet_w
            undet_y = new_df.loc[i, 'y']
            undet_h = new_df.loc[i, 'h']
            undet_y2 = undet_y + undet_h

            xmin = 1000000
            xmax = 0
            flag = -1
            for j in det_df_copy.index:
                det_x = det_df_copy.loc[j, 'x']
                det_x2 = det_df_copy.loc[j, 'x2']
                det_w = det_df_copy.loc[j, 'w']
                name = det_df_copy.loc[j, 'supercategory']
                if name not in ['table']:
                    boo = self.is_overlap_check_along_rows(undet_x, undet_x2, undet_w, det_x, det_x2, det_w)
                    if boo:
                        _, per = boo
                        if per > 0.3:
                            xmin = min(xmin, min(undet_x, det_x))
                            xmax = max(max(undet_x2, det_x2), xmax)
                            flag = 1

                            # we now try to confirm that x-min & x-max is within the range of columns
                            for col_det_ind in col_det.index:
                                label = col_det.loc[col_det_ind, 'name']
                                if label in ['column']:
                                    cox = col_det.loc[col_det_ind, 'x']
                                    coy = col_det.loc[col_det_ind, 'y']
                                    cox2 = col_det.loc[col_det_ind, 'x2']
                                    coy2 = col_det.loc[col_det_ind, 'y2']
                                    # overlap with column
                                    boocol, areacol, percol = self.isRectangleOverlap(R1=[cox, undet_y, cox2, undet_y2], R2=[undet_x, undet_y, undet_x2, undet_y2])
                                    if boocol is True and percol > 0.5:
                                        xmin = max(xmin, cox)   # this is delibrately set as Max
                                        xmax = min(xmax, cox2)  # this is delibrately kept 'min' so that word-x2 does not exceed the x2 of column

            if flag > 0:
                # overlap with Exsisting Detection was found
                undet_frame = undet_frame.append({'id': ctr, 'x': xmin, 'y': undet_y, 'w': xmax - xmin, 'h': undet_h, 'x2': xmax, 'y2': undet_y + undet_h, 'label': 'cell', 'conf': 2, 'category_id': 4, 'name': 'cell', 'supercategory': 'sentence'}, ignore_index=True)
                undet_frame.drop_duplicates(inplace=True)

            if flag < 0:
                # overlap with any previous detection was not found and hence Meta information is updated
                undet_frame = undet_frame.append({'id': ctr, 'x': undet_x, 'y': undet_y, 'w': undet_w, 'h': undet_h, 'x2': undet_x2, 'y2': undet_y + undet_h, 'label': 'cell', 'conf': 2, 'category_id': 4, 'name': 'cell', 'supercategory': 'sentence'}, ignore_index=True)
                undet_frame.drop_duplicates(inplace=True)

        # we now use our function to see if there is any overlap between
        # the undetected detections formed
        undet_frame.drop_duplicates(inplace=True)
        undet_frame.sort_values(by=['x'], inplace=True)
        undet_frame.reset_index(drop=True, inplace=True)
        undet_frame['id'] = undet_frame.index + ctr
        undet_frame_ = self.overlap_of_detections_check(undet_frame)

        if len(undet_frame) > 0:
            undet_frame_2 = self.overlap_det_removal_after_check(undet_frame_, id_ctr=undet_frame.loc[len(undet_frame) - 1, 'id'] + 1)
            undet_frame_2['id'] = -1
            undet_frame_2.drop_duplicates(inplace=True)
            undet_frame_2.reset_index(drop=True, inplace=True)
            undet_frame_2['id'] = undet_frame_2.index + ctr

            # Merging the unde frame with our prev detections
            det_df_copy = det_df_copy.append(undet_frame_2, ignore_index=True)
            det_df_copy.drop_duplicates(inplace=True)
            det_df_copy.reset_index(drop=True, inplace=True)
            det_df_copy['id'] = det_df_copy.index

            # again performing the possibl overlap between undetected anootations made + the detections by model
            det_df_copy = self.overlap_of_detections_check(det_df_copy)
            det_df_copy_ = self.overlap_det_removal_after_check(det_df_copy, id_ctr=det_df_copy.loc[len(det_df_copy) - 1, 'id'] + 1)

            det_df_copy_ = det_df_copy_.loc[det_df_copy_['h'] > 5]              # Making sure all detections atleast have a min height of 5 pixel
            det_df_copy_.sort_values(by=['y', 'x'], inplace=True)
            det_df_copy_.drop_duplicates(inplace=True)
            det_df_copy_.reset_index(drop=True, inplace=True)

            # rerunning cell overlap section again to incorporate new detections
            input_df = self.df_to_table_df_v2(input_df=self.df_raw, det_df=det_df_copy_)

            # NOTE :
            # new_df : is Meta info dataframe, which has information of all words Inside Table but Cell was not detected
            # det_df_copy : detection information of cells inside the table Only + undetected annotations created
            return input_df, det_df_copy_, new_df
        else:
            self.db_status = 'ML_ready'
            # rerunning cell overlap section again to incorporate new detections ie
            # removing annotaton that do not belong to table
            input_df = self.df_to_table_df_v2(input_df=self.df_raw, det_df=det_df_copy)
            return input_df, det_df_copy, new_df

    def height_correction_after_undetected_annotation(self, df, ids_to_work_on):
        '''
        Here we make sure that Annotations that were manually created have a strict height of the word only
        '''
        ids = list(ids_to_work_on.index)
        for id in ids:
            df.loc[id, 'c_y'] = df.loc[id, 'y']
            df.loc[id, 'c_h'] = df.loc[id, 'h']

        # now we prepape lookup_detections again
        new_lookup = pd.DataFrame(columns=self.lookup_detections_df.columns)
        df_lok_n = df[['c_x', 'c_y', 'c_w', 'c_h']]
        df_lok_n = df_lok_n.drop_duplicates()
        for ind in df_lok_n.index:
            x = df_lok_n.loc[ind, 'c_x']
            y = df_lok_n.loc[ind, 'c_y']
            w = df_lok_n.loc[ind, 'c_w']
            h = df_lok_n.loc[ind, 'c_h']
            x2 = x+w
            y2 = y+h
            if x > 0:
                new_lookup = new_lookup.append({'x':x, 'y':y, 'w':w, 'h':h,'x2':x2, 'y2':y2, 'label': 'cell',
                                                'conf':1, 'category_id':4, 'name':'cell', 'supercategory':'sentence'},
                                               ignore_index=True)

        df_tab = self.lookup_detections_df.loc[self.lookup_detections_df['supercategory']=='table']
        new_lookup = new_lookup.append(df_tab, ignore_index=True)

        # restting new
        new_lookup = new_lookup.drop_duplicates()
        new_lookup.reset_index(drop=True, inplace=True)

        return df, new_lookup

    def df_to_table_df_v2(self, input_df, det_df):
        '''
        this function takes as input :
        Params :
            input_df        : the dataframe extracted from pymupdf with unfilled table_cell_info
            det_df          : dataframe created which has detection information done by model

        Return :
            df0             : updated input_df with overlap information of table or cell
        '''
        df0 = input_df.copy()
        for j in range(len(input_df)):
            v_x = df0.loc[j, 'x']
            v_y = df0.loc[j, 'y']
            v_w = df0.loc[j, 'w']
            v_h = df0.loc[j, 'h']
            v_x2 = v_x + v_w
            v_y2 = v_y + v_h
            overlap_index = []

            for i in range(len(det_df)):
                d_x = det_df.loc[i, 'x']
                d_y = det_df.loc[i, 'y']
                d_x2 = det_df.loc[i, 'x2']
                d_y2 = det_df.loc[i, 'y2']
                label = det_df.loc[i, 'label']
                cell_id = det_df.loc[i, 'id']

                if label not in ['semi_grided_table', 'grided_table',
                                 'gridless_table', 'key_value']:
                    boo, area, perc = self.isRectangleOverlap([v_x, v_y, v_x2, v_y2], [d_x, d_y, d_x2, d_y2])
                    if boo == True and perc > 0.4:
                        overlap_index.append([i, cell_id, area, perc])
                        df0.loc[j, 'is_in_cell'] = 1
                        df0.loc[j, 'cell_id'] = cell_id
                        df0.loc[j, 'cell_type'] = label
                        df0.loc[j, 'c_x'] = d_x
                        df0.loc[j, 'c_y'] = d_y
                        df0.loc[j, 'c_w'] = d_x2 - d_x
                        df0.loc[j, 'c_h'] = d_y2 - d_y
                elif label in ['semi_grided_table', 'grided_table', 'gridless_table', 'key_value']:
                    boo, area, perc = self.isRectangleOverlap([v_x, v_y, v_x2, v_y2], [d_x, d_y, d_x2, d_y2])
                    if boo == True and perc > 0.4:
                        df0.loc[j, 'is_in_table'] = 1
                        df0.loc[j, 'table_id'] = cell_id
                        df0.loc[j, 'table_type'] = label
                        df0.loc[j, 't_x'] = d_x
                        df0.loc[j, 't_y'] = d_y
                        df0.loc[j, 't_w'] = d_x2 - d_x
                        df0.loc[j, 't_h'] = d_y2 - d_y

                df0.loc[j, 'cell_overlap'] = str(overlap_index)

        return df0

    def is_overlap_check_along_rows(self, ax1, ax2, aw, bx1, bx2, bw, perc =0.07):
        '''
        This Function checks if there is an overlap between
        x - coordinates of cells of boxes, irrespective of the y- coordinate
        '''
        if aw < bw:
            if ((ax1 >= bx1) and (ax2 > bx1)) and ((ax1 < bx2) and (ax2 <= bx2)):
                return True, 1
        if bw < aw:
            if ((bx1 >= ax1) and (bx2 > ax1)) and ((bx1 < ax2) and (bx2 <= ax2)):
                return True, 1
        if ax1 <= bx1:
            if (bx1 >= ax1) and (bx2 > ax1) and (bx1 < ax2):  # ie bx1 lies between ax1 and ax2
                per = (ax2 - bx1) / (ax2 - ax1)
                if per > perc:
                    return True, per
        if bx1 <= ax1:
            if (ax1 >= bx1) and (ax2 > bx1) and (ax1 < bx2):  # ie bx1 lies between ax1 and ax2
                per = (bx2 - ax1) / (bx2 - bx1)
                if per > perc:
                    return True, per

    def isRectangleOverlap(self, R1, R2):
        '''
        This function detects if two rectangles overlap
        '''

        if (R1[0] >= R2[2]) or (R1[2] <= R2[0]) or (R1[3] <= R2[1]) or (R1[1] >= R2[3]):
            return False, 0, 0
        else:
            # intersection along X - axis
            dx = min(R1[2], R2[2]) - max(R1[0], R2[0])
            dy = min(R1[3], R2[3]) - max(R1[1], R2[1])

            area = dx * dy

            wR1 = R1[2] - R1[0]
            wR2 = R2[2] - R2[0]

            if wR1 > wR2:
                # Rectangle R2 is small
                area_small_R = wR2 * (R2[3] - R2[1])
            else:
                area_small_R = wR1 * (R1[3] - R1[1])

            # to avoid ZerodivisionError
            if area_small_R > 1:
                per = round(area / area_small_R, 2)
                # print(area, per)
                return True, area, per
            else:
                return False, 0, 0

    def line_num_correction(self, src_df):

        df_n = src_df.sort_values(['y', 'word_n', 'page_number'], axis=0)
        ctr = 0
        zn = []
        val = None
        for iter, ind in enumerate(df_n.index):  # df_n is sorted on y

            if iter == 0:
                val = df_n.loc[ind, 'y']
                zn.append(ctr)
            else:
                if (df_n.loc[ind, 'y'] in range(val, val+5)) or (df_n.loc[ind, 'y'] in range(val-5, val)):
                    zn.append(ctr)
                else:
                    ctr = ctr + 1
                    val = df_n.loc[ind, 'y']
                    zn.append(ctr)
        df_n['line_n'] = zn
        df_n.sort_values(by=['y', 'x'], inplace=True)
        df_n.reset_index(drop=True, inplace=True)
        return df_n

    def word_num_corr(self, src_df):
        zdf = src_df.sort_values(['line_n', 'x', 'page_number'], axis=0)
        ctr = 0
        wn = []
        prev_ind = None
        for iter, ind in enumerate(zdf.index):
            if iter == 0 and zdf.loc[ind, 'line_n'] == 0:
                prev_ind = ind
                wn.append(ctr)
            else:
                if zdf.loc[ind, 'line_n'] == zdf.loc[prev_ind, 'line_n']:
                    ctr += 1
                    wn.append(ctr)
                else:
                    ctr = 0
                    wn.append(ctr)
                prev_ind = ind

        zdf['word_n'] = wn
        return zdf

    def block_corr_intrim_1(self, src_df):
        """
        This function takes src_dataframe after line correction and word_correction
        and identifies the line having maximum number of distinct cell_x coordinates detected.
        :param src_df: Dataframe after Word_correction
        :return: line number having max distinct cell _x coordinates
        """
        max_len_line = 0
        dist_cells_in_line = 0
        ln = None
        for i in src_df['line_n'].unique():
            zt = src_df.loc[src_df['line_n'] == i]
            t = len(zt['c_x'].value_counts())
            min_x_point = min(zt['c_x'])
            max_x_point = max(zt['c_x']) + zt.loc[zt['c_x'].idxmax(), 'c_w']
            if ((max_x_point - min_x_point) > max_len_line) and t >= dist_cells_in_line:
                max_len_line = max_x_point - min_x_point
                dist_cells_in_line = t
                ln = i
        return ln

    def block_corr_intrim_2(self, src_df):
        '''
        z_a is the reference frame we produce on which overlapping will be checked further
        '''

        line = self.block_corr_intrim_1(src_df)

        z_a = src_df.loc[src_df['line_n'] == line]
        z_a = z_a[['block_n', 'c_x', 'c_y', 'c_w', 'c_h']]
        z_a = z_a.sort_values(['c_x'])

        # correct block values of this reference frame
        uni = list(z_a['c_x'].unique())
        uni.sort()
        blk = 0
        for e in uni:
            for ind in z_a.index:
                x = z_a.loc[ind, 'c_x']
                if e == x:
                    z_a.loc[ind, 'block_n'] = blk
            blk = blk + 1
        z_a = z_a.drop_duplicates()
        return z_a

    def block_corr_final(self, src_df):
        '''
        src_df is the datframe where we will manipulate the block numbers.
        ref_df is one we get from block_correction_intrim_2 based on the line identified for
        creating blocks
        '''

        ref_df = self.block_corr_intrim_2(src_df)

        for ind in src_df.index:
            c_x1 = src_df.loc[ind, 'c_x']
            c_x2 = src_df.loc[ind, 'c_x'] + src_df.loc[ind, 'c_w']
            c_w = src_df.loc[ind, 'c_w']
            for rind in ref_df.index:
                r_x1 = ref_df.loc[rind, 'c_x']
                r_x2 = ref_df.loc[rind, 'c_x'] + ref_df.loc[rind, 'c_w']
                r_w = ref_df.loc[rind, 'c_w']
                if self.is_overlap_check_along_rows(c_x1, c_x2, c_w, r_x1, r_x2, r_w):
                    src_df.loc[ind, 'block_n'] = ref_df.loc[rind, 'block_n']
        return src_df

    def identify_cell_ids_with_multiline_v1_old(self, src):
        '''
        This function handles the multilines deteccted by the model.
        if multiline is not detected this function cannot handle it
        Return :
                Model returns the orignal_frame where multiline is not concatated &
                dataframe where multiline is merged.
        '''
        df_copy = src.copy()
        cell_id_uniq = list(src['cell_id'].unique())
        lst = []
        for id_ in cell_id_uniq:
            temp = src.loc[src['cell_id'] == id_]
            temp = temp.reset_index(drop=False)
            for i in range(len(temp) - 1):
                # condition to see if multiline has been detected by model
                if temp.loc[i, 'c_x'] == temp.loc[i + 1, 'c_x'] and temp.loc[i, 'c_y'] == temp.loc[i + 1, 'c_y']:
                    if temp.loc[i, 'line_n'] != temp.loc[i + 1, 'line_n']:
                        lst.append(id_)

        for ids in lst:
            temp = src.loc[src['cell_id'] == ids]
            temp = temp.reset_index(drop=False)
            max_index = max(temp['index'])
            text = ''
            for i in range(len(temp)):
                ind = temp.loc[i, 'index']
                text = text + temp.loc[i, 'text'] + ' '
                src.loc[ind, 'text'] = ''
            src.loc[max_index, 'text'] = text
            src.drop(src.index[src['text'] == ''], inplace=True)

        # re arranging the line no after dropping empty text lines
        src = self.line_num_correction(src)
        return df_copy, src

    def identify_cell_ids_with_multiline(self, src):
        '''
        Modification made :
            1. Lst is now a Dict and contains cell id and rows index to merge
            2. append multiline value to upper cell
        This function handles the multilines deteccted by the model.
        if multiline is not detected this function cannot handle it
        Return :
                Model returns the orignal_frame where multiline is not concatated &
                dataframe where multiline is merged.
        '''
        df_copy = src.copy()
        cell_id_uniq = list(src['cell_id'].unique())

        # 'lst' ----> is basically a dictionay with 'cell_id' as KEY & 'idexes' as VALUE
        # 'lst' is populated only when there is change in line numbers in the temp frame, indicating multi-line present
        lst = {}
        for id_ in cell_id_uniq:
            temp_sup = src.loc[src['cell_id'] == id_]
            temp_sup = temp_sup.reset_index(drop=False)
            y_axis_uniq = list(temp_sup['c_y'].unique())
            for y in y_axis_uniq:
                temp = temp_sup.loc[temp_sup['c_y'] == y]
                temp = temp.reset_index(drop=False)
                for i in range(len(temp) - 1):
                    # condition to see if multiline has been detected by model
                    if temp.loc[i, 'c_x'] == temp.loc[i + 1, 'c_x'] and temp.loc[i, 'c_y'] == temp.loc[i + 1, 'c_y']:
                        if temp.loc[i, 'line_n'] != temp.loc[i + 1, 'line_n']:
                            lst.update({id_: list(temp['index'])})

        # olap_dic_  -----> is a dictionary where 'cell_id' is KEY & cellIDs with which there is y-axis overlaps is VALUE
        olap_dic_ = self.identify_overlap_det_y_axis_multiline(df_copy, cell_id_uniq, lst)
        for ids in lst.keys():
            temp = src.loc[lst[ids]]
            temp = temp.sort_values(by=['y', 'x'])
            temp = temp.reset_index(drop=False)
            max_index = max(temp['index'])
            min_index = min(temp['index'])
            # get line number of this min_index identified
            line_no = src.loc[min_index, 'line_n']
            y_val = src.loc[min_index, 'y']
            text = ''
            for i in range(len(temp)):
                ind = temp.loc[i, 'index']
                text = text + temp.loc[i, 'text'] + ' '
                src.loc[ind, 'text'] = ''
            src.loc[min_index, 'text'] = text

            # trying to change 'line_n' & 'y' of all cells that y-axis overlaps wit 'ids' here
            if len(olap_dic_) > 0:
                olap_lst_ = olap_dic_[ids]
                for cid in olap_lst_:
                    if cid in lst.keys():
                        # checking if the 'cid' is a multiline cell and if so must exsist in 'lst_' as a KEY
                        tmp_change_frame = src.loc[src['cell_id'] == cid]
                        tmp_change_frame = tmp_change_frame.sort_values(by=['y', 'x'])
                        # update 25-Mar-2022  : it is observed that y value of all must not be changed
                        # hence only first text in tmp_change_frame will have y-val modified
                        for ind in tmp_change_frame.index:
                            src.loc[ind, 'line_n'] = line_no
                            src.loc[ind, 'y'] = y_val               #
                            break
                    else:
                        # if the 'cid' is NOT a multiline cell and if so must exsist in 'lst_' as a KEY
                        tmp_change_frame = src.loc[src['cell_id'] == cid]
                        tmp_change_frame = tmp_change_frame.sort_values(by=['y', 'x'])
                        # update 26-Mar-2022  : it is observed that y value of all must be changed if cid is not multiline
                        for ind in tmp_change_frame.index:
                            src.loc[ind, 'line_n'] = line_no
                            src.loc[ind, 'y'] = y_val

            src.drop(src.index[src['text'] == ''], inplace=True)

        # re arranging the line no after dropping empty text lines
        src = self.line_num_correction(src)
        return df_copy, src

    def multiline_correction_final(self, input_df):
        src = input_df.copy()
        src.sort_values(by=['c_y', 'c_x'], ascending=True, inplace=True)
        for ind in src.index:
            com_cy = src.loc[ind, 'c_y']
            com_cy2 = src.loc[ind, 'c_h'] + com_cy
            for jind in src.index:
                if ind != jind:
                    sub_cy = src.loc[jind, 'c_y']
                    sub_cy2 = src.loc[jind, 'c_h'] + com_cy
                    if com_cy <= sub_cy <= com_cy2:
                        src.loc[ind, 'line_n'] = min(src.loc[ind, 'line_n'], src.loc[jind, 'line_n'])
                        src.loc[jind, 'line_n'] = min(src.loc[ind, 'line_n'], src.loc[jind, 'line_n'])

        src.sort_values(['c_x', 'c_y'])
        src.reset_index(drop=True, inplace=True)
        return src

    def handle_multiline_with_rule(self, df):
        """
        the working is based on rule based model
        """
        src = df.copy()

        # 1.a : storing column names in col
        # col = src.columns.sort_values(ascending=True).to_list()
        # 1.b : replacing all empty cells with nan
        src.replace(r'\t', '', regex=True, inplace=True)
        src.replace(r'^\s*$', np.nan, regex=True, inplace=True)
        # 1.c : dropping all rows where all values are NaN
        src = src.dropna(how='all', axis=0)
        # 1.d : droping all columns where all values are NaN
        src = src.dropna(how='all', axis=1)
        col = src.columns.sort_values(ascending=True).to_list()
        # STEP 2 :
        # -------------- Identifying datatype of Each Columns --------------

        col_type = {}
        for c in col:
            ct_num = 0
            ct_othr = 0
            for ind in src.index:
                val = src.loc[ind, c]
                if not pd.isna(val):
                    val = val.strip().replace(',', '')
                    val = val.strip().replace('$', '')
                    # val = val.strip().replace('.', '')
                    try:
                        val = float(val)
                        src.loc[ind, c] = val
                        ct_num += 1
                    except:
                        ct_othr += 1
                else:
                    ct_othr += 1
            if ct_num > ct_othr:
                col_type.update({c: 'numeric'})
            else:
                col_type.update({c: 'other'})
        # ---------------------------------------------------------------------
        # STEP 3:  Identifying All Index where NaN is present
        rows_with_nan = [index for index, row in src.iterrows() if row.isnull().any()]
        rows_with_no_nan = [index for index, row in src.iterrows() if not row.isnull().any()]
        rows_with_nan_ = rows_with_nan.copy()
        rows_with_nan_n_num_n_str = []  # indexes where we have nan + number + strings

        # check index identified as nan has a numerical value
        for ind in rows_with_nan_:
            for c in col:
                val = src.loc[ind, c]
                if not pd.isna(val):  # if value is not nan
                    if not isinstance(val, (int, float, complex)):
                        try:
                            val = val.strip().replace(',', '')
                            val = float(val)
                            rows_with_nan.remove(ind)
                            rows_with_nan_n_num_n_str.append(ind)
                            src.loc[ind, c] = val  # changing the type in source dtaframe to numeric if possible
                        except:
                            pass
                    else:
                        rows_with_nan_n_num_n_str.append(ind)

        # STEP 4 :
        rows_with_nan_n_num_only = rows_with_nan_n_num_n_str.copy()
        rows_with_nan_n_num_only_not_in_numeric_col = []
        # check of index which have numbers, how many have a string too
        for e in rows_with_nan_n_num_n_str:
            total_na_in_row = src.loc[e, :].isna().sum()
            if total_na_in_row == len(col) - 1:
                for c in col:
                    val = src.loc[e, c]
                    if not pd.isna(val):
                        if isinstance(val, (int, float, complex)):
                            type_of_col = col_type[c]
                            if type_of_col == 'other':
                                rows_with_nan.append(e)
                                rows_with_nan_n_num_only_not_in_numeric_col.append(e)
                                rows_with_nan_n_num_only.remove(e)
            else:
                rows_with_nan.append(e)
                rows_with_nan_n_num_only.remove(e)
        rows_with_nan.sort()
        rows_with_nan_n_num_n_str = list((set(rows_with_nan_n_num_n_str) - set(rows_with_nan_n_num_only_not_in_numeric_col)) - set(rows_with_nan_n_num_only))
        rows_with_nan_n_num_n_str = list(set(rows_with_nan_n_num_n_str))
        del [rows_with_nan_]

        # --------------------------- ------------------------------------
        # STEP 6 : rule based merging
        # 1. all values in row is nan except 1, and row following that is also nan, merge up
        # 2.  all values in row is nan except 1, and row above it is in rows_with_nan
        rows_with_nan = list(set(rows_with_nan))
        rows_with_nan_processed = []  # to keep track of processed indexes

        # processing rows having a Nan Value
        for ele in rows_with_nan:
            total_na_in_row = src.loc[ele, :].isna().sum()
            # check if complete row is nan or not
            if total_na_in_row != len(col):
                # check if above row is  not in rows_with_nan_n_num
                if int(ele) - 1 not in rows_with_nan_n_num_only:
                    # check if all column except 1 is nan or not
                    if total_na_in_row == len(col) - 1 and ele not in rows_with_nan_n_num_only:
                        # ------------- UPDATE 27-mAR-2022, check if current index is not in -------------------
                        # ----rows_with_nan_n_num_only & rows_with_nan_n_num_only_not_in_numeric_col ---------
                        if int(ele) not in rows_with_nan_n_num_only and int(ele) not in rows_with_nan_n_num_only_not_in_numeric_col:
                            # check if next ind is in rows with nan
                            if int(ele) + 1 in rows_with_nan:
                                for c in col:
                                    val = src.loc[ele, c]
                                    if not pd.isna(val):
                                        column = col.copy()
                                        column.remove(c)
                                        index = ele
                                        # reconfirming the situation
                                        cond1 = total_na_in_row == len(col) - 1
                                        cond2 = src.loc[index + 1, c]
                                        try:
                                            is_char_lower = cond2[0].islower()
                                        except:
                                            is_char_lower = False
                                        if cond1 and pd.isna(cond2):
                                            # under this condition we check val of col 'c' at ind ele+1 is nan
                                            # so we merge all column except 'c' UP
                                            for tc in column:
                                                # we are merging below row UP, but advisable to convert nan-->''
                                                # and concatenate, than just replacing the values
                                                curr_val = src.loc[index, tc]
                                                next_val = src.loc[index + 1, tc]
                                                if pd.isna(curr_val):
                                                    curr_val = ''
                                                if pd.isna(next_val):
                                                    next_val = ''

                                                # since merging up
                                                new_val = str(curr_val) + ' ' + str(next_val)
                                                src.loc[index, tc] = new_val
                                                src.loc[index + 1, tc] = np.nan
                                            rows_with_nan_processed.append(ele)
                                            # rows_with_nan_processed.append(ele + 1)

                                        if (cond1 == True) and (is_char_lower == True):
                                            # under this condition we check value of col 'c' at ind ele61 is NOT nan
                                            # so we merge ALL column up
                                            for tc in col:
                                                curr_val = src.loc[index, tc]
                                                next_val = src.loc[index + 1, tc]
                                                if pd.isna(curr_val):
                                                    curr_val = ''
                                                if pd.isna(next_val):
                                                    next_val = ''
                                                # Merging UP
                                                new_val = str(curr_val) + ' ' + str(next_val)
                                                src.loc[index, tc] = new_val
                                                src.loc[index + 1, tc] = np.nan
                                            rows_with_nan_processed.append(ele)
                                            # rows_with_nan_processed.append(ele + 1)

                if int(ele) - 1 not in rows_with_nan_n_num_only and ele not in rows_with_nan_n_num_n_str and ele not in rows_with_nan_processed and ele not in rows_with_nan_n_num_only:
                    # all such cells will be merged up
                    isUpFlag = False
                    text_to_put_if_isUpFlag = []
                    if ele != 0:  # if index or ele is 0 there is no chance of merging it above
                        for tc in col:
                            try:
                                text1 = src.loc[ele - 1, tc]
                            except:
                                text1 = ''
                            if pd.isna(text1):
                                text1 = ''
                            else:
                                text1 = str(text1).strip()

                            text2 = src.loc[ele, tc]
                            if pd.isna(text2):
                                text2 = ''
                            else:
                                text2 = str(text2).strip()
                                checkIsUpper = text2[0].isupper() or text2[0] in ('*', '**')
                                checkSpecialChar_to_AvoidMerge = ('---' in text2) or ('===' in text2)  or (text2[0] in ['a', 'b', 'c', 'd', 'e']) or (text2[0:3] in ['(a)', '(b)', '(c)', '(d)', '(e)'])
                                if isUpFlag is False:
                                    isUpFlag = checkIsUpper or checkSpecialChar_to_AvoidMerge
                            text_to_put_if_isUpFlag.append(text1 + ' ' + text2)
                            # we check here if line below starts witha capital letter or not
                            # bcoz most likely if the text starts witha capital letter it must not be merged
                        if not isUpFlag and ele - 1 in src.index:
                            for ier, tc in enumerate(col):
                                src.loc[ele - 1, tc] = text_to_put_if_isUpFlag[ier]
                                src.loc[ele, tc] = np.nan
                            rows_with_nan_processed.append(ele)
                        else:
                            # say for a column with nan we merged but be actually should not have
                            # hence we remove it again from list of nan processed
                            try:
                                rows_with_nan_processed.remove(ele)
                            except:
                                pass
                if int(ele) - 1 not in rows_with_nan_n_num_only and ele not in rows_with_nan_processed:
                    if ele in rows_with_nan_n_num_n_str:
                        if int(ele) - 1 in rows_with_nan and int(ele) - 2 in rows_with_nan and int(ele) - 1 in rows_with_nan_processed:
                            # text will be merged two levels up if all is nan at ele-1 index
                            if int(ele)-1 in src.index:
                                total_na_in_row_tmp = src.loc[ele-1, :].isna().sum()
                                if total_na_in_row_tmp == len(col):
                                    for tc in col:
                                        text1 = src.loc[ele - 2, tc]
                                        text2 = src.loc[ele, tc]
                                        if pd.isna(text1):
                                            text1 = ''
                                        if pd.isna(text2):
                                            text2 = ''
                                        src.loc[ele - 2, tc] = str(text1) + str(text2)
                                        src.loc[ele, tc] = np.nan
                                    rows_with_nan_processed.append(ele)
                if ele in rows_with_nan_n_num_n_str and ele not in rows_with_nan_processed:
                    # check if Row Below has a Nan AND below row is not in nan_num_str ie. there is no number below
                    if ele + 1 not in rows_with_nan_n_num_n_str and ele+1 in src.index:
                        # In these scenarios row below will be merged up if row below starts with lower case
                        curr_val_lst = []
                        next_val_lst = []
                        merge_flag = False
                        for c in col:
                            curr_val = src.loc[ele, c]
                            if pd.isna(curr_val):
                                curr_val = ''
                            curr_val = str(curr_val).strip()
                            curr_val_lst.append(curr_val)
                            next_val = src.loc[ele + 1, c]
                            if pd.isna(next_val):
                                next_val = ''
                            next_val = str(next_val).strip()
                            next_val_lst.append(next_val)

                            if len(str(next_val)) > 0 and str(next_val)[0].islower():
                                merge_flag = True

                        # observed and updated 22-Mar-2022
                        # if '*' or ':' in any word is present in the below line
                        # it indicates a new heading , hence merge is to be avoided
                        tmp = [False for e in next_val_lst if (':' in e) or ('*' in e) or ('-----' in e) or ('===') in e]
                        if False in tmp:
                            merge_flag = False
                        if merge_flag is True:
                            for c in col:
                                src.loc[ele, c] = str(curr_val_lst[c]) + ' ' + str(next_val_lst[c])
                                src.loc[ele + 1, c] = np.nan
                                rows_with_nan_processed.append(ele)
                else:
                    # rows_with_nan_processed.append(ele)
                    pass
            else:
                src.drop(ele, inplace=True)
                rows_with_nan_processed.append(ele)

        # processing rows not having any Nan Value
        # case 1 : complete row is filled but the row above has nan values & current row starts with
        # lower case
        for ele in rows_with_no_nan:
            # check if ind is  not zero ie ele-1 exists and the row above must have-had NaN
            if ele - 1 in src.index and ele - 1 in rows_with_nan:
                merge_flag = False
                col_copy = col.copy()
                for c in col:
                    val = src.loc[ele - 1, c]
                    # check if value in above col is nan
                    if pd.isna(val):
                        col_copy.remove(c)

                # now considering column/columns where above value is not nan
                for c in col_copy:
                    curr_val = src.loc[ele, c]
                    prev_val = src.loc[ele - 1, c]
                    # make sure that neither current value nor prev value is of type--> 'numeric'
                    if not isinstance(curr_val, (int, float, complex)) and not isinstance(prev_val,(int, float, complex)):
                        # check if text below starts with a lower case letter
                        if curr_val[0].islower():
                            # then we merge the above cell down
                            merge_flag = True
                if merge_flag is True:
                    for c in col:
                        curr_val = src.loc[ele, c]
                        prev_val = src.loc[ele - 1, c]
                        if pd.isna(prev_val):
                            prev_val = ''

                        new_curr_val = str(prev_val) + ' ' + str(curr_val)
                        src.loc[ele, c] = new_curr_val
                        # changing above cell val to NaN
                        src.loc[ele - 1, c] = np.nan
        src = src.dropna(how='all', axis=0)
        src.reset_index(drop=True, inplace=True)

        # checking if top row has just (a), (b), (c), (d) uptill (f)
        cols = src.columns.sort_values(ascending=True).to_list()
        for cs in cols:
            try:
                val = src.loc[0, cs].strip(' ')
            except:
                val = src.loc[0, cs]
            if val in ['a', 'b', 'c', 'd', 'e', '(a)', '(b)', '(c)', '(d)', '(e)', '(f)']:
                nxt_val = src.loc[1, cs]
                try:
                    nxt_val = nxt_val.strip()
                except:
                    pass
                if nxt_val not in ('*', '**'):
                    if pd.isna(nxt_val):
                        nxt_val = ''
                    try:
                        nxt_val = nxt_val.strip(' ')
                    except:
                        nxt_val = nxt_val
                    src.loc[0, cs] = str(val) + ' ' + str(nxt_val)
                    src.loc[1, cs] = np.nan
        src = src.dropna(how='all', axis=0)
        src.reset_index(drop=True, inplace=True)

        return src

    def formatting_table_to_excel_like(self, src_frame, page_no):
        """
        this function takes input as the Source frame and path where to save the excel file generated
        :param src_frame: Source Dataframe
        :return: dataframe which can be written
        """

        if len(src_frame) > 0:

            src_frame = src_frame.sort_values(['c_x', 'c_y'])
            src_frame.reset_index(drop=True, inplace=True)

            no_of_lines = max(src_frame['line_n']) + 1
            max_col = max(src_frame['block_n']) + 1
            df_ = pd.DataFrame(columns=range(1, max_col))

            for line_number in range(1, no_of_lines + 1):
                v = dict((el, '') for el in range(max_col))
                temp_frame = src_frame.loc[src_frame['line_n'] == line_number - 1]
                temp_frame = temp_frame.reset_index(drop=True)
                for r in range(len(temp_frame)):
                    tmp = temp_frame.loc[r, 'block_n']
                    v[tmp] = v[tmp] + temp_frame.loc[r, 'text'] + ' '

                df_ = df_.append(v, ignore_index=True)
            df_ = df_.reindex(columns=sorted(df_.columns))
            return df_
        else:
            print("ERROR : TABLE not detected for page no ", page_no)
            return pd.DataFrame(columns=[1, 2, 3])

    def write_excel_after_multiline_handling(self, src_frame, save_path, read_path, page_no, save_excel):

        # ---------------------renaming columns to correct sequence if any column was dropped ---
        new_col = {}
        col = list(src_frame.columns)
        for iter, ind in enumerate(col):
            new_col.update({ind: iter})
        # src_frame = src_frame.rename(columns=new_col)
        src_frame = src_frame.reindex(columns=sorted(src_frame.columns))
        # ----------------------------------------------------------------------------------------
        if save_excel:
            if os.path.exists(read_path):
                myworkbook = load_workbook(read_path)
                # os.remove(read_path)
            else:
                myworkbook = Workbook()
            if 'Sheet' in myworkbook.sheetnames:
                myworkbook.remove(myworkbook['Sheet'])
            if str(page_no) in myworkbook.sheetnames:
                myworkbook.remove(myworkbook[str(page_no)])
            worksheet = myworkbook.create_sheet(str(page_no))
            worksheet.title = str(page_no)
            for i in range(len(src_frame)):
                for c in src_frame.columns:
                    val = src_frame.loc[i, c]
                    if pd.isna(val):
                        val=''
                    else:
                        try:
                            val = val.strip()
                        except:
                            pass

                    worksheet.cell(row=i + 1, column=c + 1, value=val)
            try:
                myworkbook.save(save_path)
                return True
            except:
                print('ERROR : Excel file did not save ')
                return False
        else:
            return False

    def column_detection_intrim_pred(self, src_img, det_df, page_n):

        # identify unique tables detected by Table Model

        # uniq_table_id = list(det_df['table_id'].unique())                 # using Metadata as input det_df
        uniq_table_id = list(det_df['category_id'].unique())                # using lookup_detection_df as det_df

        df_col_header = pd.DataFrame(columns=['x', 'y', 'w', 'h', 'x2', 'y2', 'name', 'conf'])

        # Crop table area to feed to col detection model
        for e in uniq_table_id:
            if e in [0, 1, 2, 3]:                                           # check if supercategory is Table
                df_temp = det_df.loc[det_df['category_id'] == e]
                # df_temp = df_temp[['t_x', 't_y', 't_w', 't_h']]
                df_temp = df_temp[['x', 'y', 'w', 'h']]
                df_temp.drop_duplicates(inplace=True)

                for ind in df_temp.index:
                    tx0 = df_temp.loc[ind, 'x']
                    ty0 = df_temp.loc[ind, 'y']
                    tx1 = df_temp.loc[ind, 'w'] + tx0
                    ty1 = df_temp.loc[ind, 'h'] + ty0

                    # crop image for table
                    crop_img = src_img[ty0:ty1, tx0:tx1]
                    # run column detection model
                    column_pred = self.get_table_v2(crop_img, page_n=page_n, object_names_list=['column'])

                    for colde in column_pred:
                        dx1, dy1, dx2, dy2, cls_label, conf = colde
                        # print(cls_label)
                        if conf > 0.9:
                            if cls_label in ['column']:
                                # df_col_header = df_col_header.append({'x': dx1 + tx0, 'y': dy1 + ty0, 'w': ((dx2 + tx0) - (dx1 + tx0)), 'h': ((dy2 + ty0) - (dy1 + ty0)), 'x2': dx2 + tx0, 'y2': dy2 + ty0, 'name': cls_label, 'conf': conf}, ignore_index=True)
                                df_col_header = df_col_header.append(
                                    {'x': dx1 + tx0, 'y': ty0+10, 'w': ((dx2 + tx0) - (dx1 + tx0)),
                                     'h': ty1-ty0-20 , 'x2': dx2 + tx0, 'y2': ty1-10,
                                     'name': cls_label, 'conf': conf}, ignore_index=True)

                            if cls_label in ['header']:
                                # df_col_header = df_col_header.append({'x': dx1 + tx0, 'y': dy1 + ty0, 'w': ((dx2 + tx0) - (dx1 + tx0)), 'h': ((dy2 + ty0) - (dy1 + ty0)), 'x2': dx2 + tx0, 'y2': dy2 + ty0, 'name': cls_label, 'conf': conf}, ignore_index=True)
                                df_col_header = df_col_header.append(
                                    {'x': tx0, 'y': dy1 + ty0, 'w': tx1- tx0,
                                     'h': ((dy2 + ty0) - (dy1 + ty0)), 'x2': tx1, 'y2': dy2 + ty0, 'name': cls_label,
                                     'conf': conf}, ignore_index=True)

        df_col_header.sort_values(['x'], inplace=True)
        df_col_header = df_col_header.reset_index(drop=True)
        return df_col_header

    def column_det_overlap_remove(self, col_head_det):
        '''
        This function removes all columns that overlap along x-asis / horizontal axis
        '''
        # Make sure columns dont overlap
        col_head_det_ = col_head_det.copy()

        header_only = col_head_det_.loc[col_head_det_['name']=='header']
        col_only = col_head_det_.loc[col_head_det_['name']=='column']
        col_only2 = col_head_det_.loc[col_head_det_['name'] == 'column']

        col_only.sort_values(by=['x2', 'name'], inplace=True)
        col_only.drop(col_only[col_only['name'] == 'header'].index, inplace=True)
        col_only.reset_index(drop=True, inplace=True)

        col_only2.sort_values(by=['x2', 'name'], inplace=True)
        col_only2.drop(col_only2[col_only2['name'] == 'header'].index, inplace=True)
        col_only2.reset_index(drop=True, inplace=True)

        for i in range(1, len(col_only)):
            current_x1 = col_only.loc[i, 'x']
            previous_x2 = col_only.loc[i - 1, 'x2']
            if previous_x2 > current_x1:
                pixels_overlap = previous_x2 - current_x1
                # previous_new = previous_x2 - int(pixels_overlap / 2)
                previous_new = min(previous_x2, current_x1)
                col_only2.loc[i, 'x'] = previous_new + 10
                col_only2.loc[i - 1, 'x2'] = previous_new
                # width value changes
                col_only2.loc[i, 'w'] = col_only2.loc[i, 'x2'] - col_only2.loc[i, 'x']
                col_only2.loc[i - 1, 'w'] = col_only2.loc[i - 1, 'x2'] - col_only2.loc[i - 1, 'x']

        lst_ = []
        for ind in col_only2.index:
            wid = col_only2.loc[ind, 'w']
            if wid < 10:
                lst_.append(ind)
        col_only2.drop(index=lst_, inplace=True)
        col_only2.reset_index(drop=True, inplace=True)

        header_only = header_only.append(col_only2, ignore_index=True)
        return header_only

    def column_remove_annotations_in_multi_columns(self, cell_det_frame, col_det_frame):
        """
        Removes cell detectos that move in multiple columns
        """

        cell_det_new = cell_det_frame.copy()
        drop_indexes = []

        for cell_det_ind in cell_det_frame.index:
            label = cell_det_frame.loc[cell_det_ind, 'name']
            if label in ['cell', 'special_cell']:
                cex = cell_det_frame.loc[cell_det_ind, 'x']
                cey = cell_det_frame.loc[cell_det_ind, 'y']
                cex2 = cell_det_frame.loc[cell_det_ind, 'x2']
                cey2 = cell_det_frame.loc[cell_det_ind, 'y2']
                ceid = cell_det_frame.loc[cell_det_ind, 'id']

                lst = []
                for col_det_ind in col_det_frame.index:
                    label = col_det_frame.loc[col_det_ind, 'name']
                    if label in ['column']:
                        cox = col_det_frame.loc[col_det_ind, 'x']
                        coy = col_det_frame.loc[col_det_ind, 'y']
                        cox2 = col_det_frame.loc[col_det_ind, 'x2']
                        coy2 = col_det_frame.loc[col_det_ind, 'y2']

                        boo, area, per = self.isRectangleOverlap(R1=[cox, cey, cox2, cey2], R2=[cex, cey, cex2, cey2])
                        if boo is True: # and per > 0.4:
                            lst.append([cox, coy, cox2, coy2])
                if len(lst) > 1:
                    drop_indexes.append(cell_det_ind)
                    # we now remove orignal detection and create new detections
                    for col in lst:
                        # Tackling scenario where cell-x is between column x-x2
                        if col[0] <= cex < col[2]:
                            # new coordinates will be
                            xmin = cex
                            xmax = col[2] - 5
                            ymin = cey
                            ymax = cey2
                            w_new = xmax - xmin
                            # it only makes sense to check if the width of new cell is > 15 pixel atleast
                            if w_new > 15:
                                cell_det_new = cell_det_new.append({'id': len(cell_det_new), 'x': xmin, 'y': ymin, 'w': w_new, 'h': ymax - ymin, 'x2': xmax, 'y2': ymax, 'label': 'cell', 'conf': 2, 'category_id': 4, 'name': 'cell','supercategory': 'sentence'}, ignore_index=True)

                        # Tackling scenario where cell-x2 is between column x-x2
                        if col[0] < cex2 <= col[2]:
                            # new coordinates will be
                            xmin = col[0] + 5
                            xmax = cex2
                            ymin = cey
                            ymax = cey2
                            w_new = xmax - xmin
                            # it only makes sense to check if the width of new cell is > 15 pixel atleast
                            if w_new > 15:
                                cell_det_new = cell_det_new.append({'id':len(cell_det_new),'x': xmin, 'y': ymin, 'w': w_new, 'h': ymax - ymin,'x2': xmax, 'y2': ymax, 'label': 'cell', 'conf': 2,'category_id': 4, 'name': 'cell','supercategory': 'sentence'},ignore_index=True)

                        # Tackling scenario where neither cell-x nor cell-x2 is between cols
                        if cex < col[0] and cex2 > col[2] and ((cex <= col[0] < cex2) and (cex < col[2] <= cex2)):
                            # new coordinates will be
                            xmin = col[0] + 5
                            xmax = col[2] - 5
                            ymin = cey
                            ymax = cey2
                            w_new = xmax - xmin
                            # it only makes sense to check if the width of new cell is > 15 pixel atleast
                            if w_new > 15:
                                cell_det_new = cell_det_new.append({'id': len(cell_det_new), 'x': xmin, 'y': ymin, 'w': w_new, 'h': ymax - ymin, 'x2': xmax, 'y2': ymax, 'label': 'cell', 'conf': 2, 'category_id': 4,'name': 'cell', 'supercategory': 'sentence'}, ignore_index=True)

        cell_det_new.drop(cell_det_new.index[drop_indexes], inplace=True)
        cell_det_new.reset_index(drop=True, inplace=True)
        cell_det_new_2 = self.overlap_of_detections_check(cell_det_new)
        cell_det_new_3 = self.overlap_det_removal_after_check(cell_det_new_2)

        return cell_det_new_3

    def cells_remove_overlap_annotations(self, col_det_df, lookup_det):

        # using column detection we get anootations of column only
        col_only = col_det_df.loc[col_det_df['name'] == 'column']
        col_only.sort_values(by=['x'], inplace=True)
        col_only.reset_index(drop=True, inplace=True)

        # in Lookup detection frame we add a column called 'col' = -1
        # and identify which column the detection will belong to
        lookup_det['col'] = -1  # --------> Lookup Detections DF
        for i in range(len(lookup_det)):
            name = lookup_det.loc[i, 'supercategory']
            if name in ['sentence']:
                ce_x = lookup_det.loc[i, 'x'] + 5
                ce_x2 = lookup_det.loc[i, 'x2']
                for jind in col_only.index:
                    co_x = col_only.loc[jind, 'x']
                    co_x2 = col_only.loc[jind, 'x2']
                    if (co_x2 > ce_x > co_x) or (co_x > ce_x2 > co_x2):
                        lookup_det.loc[i, 'col'] = jind

        lookup_copy = lookup_det.copy()
        cell_det = lookup_det.loc[lookup_det['name'] == 'cell']
        cell_det = cell_det.loc[lookup_det['col'] > -1]

        unique_col = cell_det['col'].unique()
        # considering all detections where column is greater that  -1
        for ucol in unique_col:
            if ucol >= 0:
                tmp = cell_det.loc[cell_det['col'] == ucol]
                tmp.sort_values(by=['y2'], inplace=True)
                tmp.reset_index(drop=False, inplace=True)
                for i in range(1, len(tmp)):
                    current_y1 = tmp.loc[i, 'y']
                    previous_y2 = tmp.loc[i - 1, 'y2']

                    cur_ind = tmp.loc[i, 'index']
                    pre_ind = tmp.loc[i - 1, 'index']

                    if previous_y2 > current_y1:
                        pixels_overlap = previous_y2 - current_y1
                        pre_new_val = min(current_y1, previous_y2) - 5
                        cur_new_val = max(current_y1, previous_y2) + 5

                        lookup_copy.loc[pre_ind, 'y2'] = pre_new_val
                        lookup_copy.loc[cur_ind, 'y'] = cur_new_val

                        lookup_copy.loc[pre_ind, 'h'] = lookup_copy.loc[pre_ind, 'y2'] - lookup_copy.loc[pre_ind, 'y']
                        lookup_copy.loc[cur_ind, 'h'] = lookup_copy.loc[cur_ind, 'y2'] - lookup_copy.loc[cur_ind, 'y']
        lookup_copy.drop(['col'], axis=1, inplace=True)

        return lookup_copy

    def identify_headers(self, src_df_, col_det_df):
        '''
        src_df      :   Dataframe with metadata and table -cell overlap
        col_det_df  :   Dataframe of detections by column Model
        '''
        src_df = src_df_.copy()
        for src_ind in src_df.index:
            src_rec = [src_df.loc[src_ind, 'x'], src_df.loc[src_ind, 'y'],
                       src_df.loc[src_ind, 'x'] + src_df.loc[src_ind, 'w'],
                       src_df.loc[src_ind, 'y'] + src_df.loc[src_ind, 'h']]
            for coldet_ind in col_det_df.index:
                label = col_det_df.loc[coldet_ind, 'name']
                if label in ['header', 'headers']:
                    coldet_rec = [col_det_df.loc[coldet_ind, 'x'], col_det_df.loc[coldet_ind, 'y'],
                                  col_det_df.loc[coldet_ind, 'x2'], col_det_df.loc[coldet_ind, 'y2']]
                    boo, area, per = self.isRectangleOverlap(coldet_rec, src_rec)
                    if boo is True and per > 0.3:
                        src_df.loc[src_ind, 'header'] = 1
        return src_df

    def column_creation_v1(self,  src_df, columdet):
        header_flag = -1
        frame = src_df.loc[src_df['header'] == 1]
        # check to see if 'HEADER' is detected
        if len(frame) > 0:                                      # HEADER is present in detection from Column model
            header_flag = 1
            frame.sort_values(['x'], inplace=True)
            frame = frame.reset_index(drop=False)
            for ind in frame.index:
                hx1 = frame.loc[ind, 'x']
                hx2 = frame.loc[ind, 'x'] + frame.loc[ind, 'w']
                for indcol in columdet.index:
                    name = columdet.loc[indcol, 'name']
                    if name not in ['header']:
                        colx1 = columdet.loc[indcol, 'x']
                        colx2 = columdet.loc[indcol, 'x2']
                        if hx1 >= colx1 and hx2 <= colx2 and frame.loc[ind, 'col'] == -1:
                            frame.loc[ind, 'col'] = indcol
                        else:
                            if hx1 < colx1 and frame.loc[ind, 'col'] == -1:
                                frame.loc[ind, 'col'] = -99
        else:                                                   # if ONLY Column was detected and no Header was found
            src_df.sort_values(['x', 'y'], inplace=True)
            src_df = src_df.reset_index(drop=True)

            columdet.sort_values(['x', 'y'], inplace=True)
            zcol_det_df = columdet.reset_index(drop=True)

            # this dict is used to populate the column index and percentage overlap between coordinates and column detection
            ovelap_lst = {}

            # using metadata here
            for i in range(len(src_df)):
                src_x = src_df.loc[i, 'x']
                src_x2 = src_x + src_df.loc[i, 'w']
                for jind in zcol_det_df.index:
                    col_x = zcol_det_df.loc[jind, 'x']
                    col_x2 = zcol_det_df.loc[jind, 'x2']

                    # check complete overlap
                    if src_x >= col_x and src_x2 <= col_x2 and src_df.loc[i, 'col'] == -1:
                        src_df.loc[i, 'col'] = jind
                        break
                    # check partial Overlap
                    else:
                        overlap = self.is_overlap_check_along_rows(ax1=col_x, ax2=col_x2, aw=col_x2 - col_x, bx1=src_x, bx2=src_x2, bw=src_x2 - src_x, perc=0.4)
                        if overlap:
                            boo, per = overlap
                            if boo is True:
                                ovelap_lst.update({jind: per})

                # if there was a partial overlap, we need to find index of column having Maximum
                # overlap with, hence following will run only if there was partial overlap
                if len(ovelap_lst) > 0:
                    keys = list(ovelap_lst.keys())  # in python 3, you'll need `list(i.keys())`
                    values = list(ovelap_lst.values())
                    max_overlap_was_with_col = keys[values.index(max(values))]
                    src_df.loc[i, 'col'] = max_overlap_was_with_col
            frame = src_df.copy()

        # this code creates the ctr for unidentified words in header
        ctr = 10
        for i in range(len(frame)):
            val = frame.loc[i, 'col']
            nextval = None
            temp = []
            if val < 0:
                temp.append(i)
                flag = -1
                for j in range(i + 1, len(frame)):
                    jval = frame.loc[j, 'col']
                    if flag < 0 and jval < 0:
                        temp.append(j)
                    else:
                        flag = 1
                        nextval = frame.loc[j, 'col']
                        break
            for k in temp:
                frame.loc[k, 'col'] = ctr
            ctr = ctr + 1
        #
        dic = {}
        unique_col = list(frame['col'].unique())
        for e in unique_col:
            temp_df = frame[frame['col'] == e]
            dic.update({e: [min(temp_df['x']), max(temp_df['x'] + temp_df['w'])]})
        new = pd.DataFrame.from_dict(dic)
        new = new.transpose()
        new = new.sort_values(by=[0])
        new = new.reset_index(drop=False)

        columdet = columdet.sort_values(by=['x'])
        columdet = columdet.reset_index(drop=False)
        for len_ind in new.index:
            index_val = new.loc[len_ind, 'index']
            len_xmin = new.loc[len_ind, 0]
            len_xmax = new.loc[len_ind, 1]

            # getting values from col detection model
            if index_val in list(columdet['index']):
                det_xmin = columdet.loc[index_val, 'x']
                det_xmax = columdet.loc[index_val, 'x2']

                xmin = min(len_xmin, det_xmin)
                xmax = max(len_xmax, det_xmax)

                #
                new.loc[len_ind, 0] = xmin
                new.loc[len_ind, 1] = xmax
                columdet.loc[index_val, 'x'] = xmin
                columdet.loc[index_val, 'x2'] = xmax
                columdet.loc[index_val, 'w'] = xmax - xmin
            else:
                xmin = len_xmin
                xmax = len_xmax
                columdet = columdet.append(
                    {'index': index_val, 'x': xmin, 'y': 0, 'w': xmax - xmin, 'h': 0, 'x2': xmax, 'y2': 0,
                     'name': 'column', 'conf': 1}, ignore_index=True)
        columdet = columdet.sort_values(by=['x'])
        columdet = columdet.reset_index(drop=True)

        lst = []

        for ind in new.index:
            old_val = new.loc[ind, 'index']
            new_val = ind
            for k in range(len(frame)):
                if frame.loc[k, 'col'] == old_val:
                    lst.append(new_val)
        frame['col'] = lst

        if header_flag < 0:
            src_df = frame.copy()

        if header_flag > 0:
            # setting new col values identified into source Fame
            for i in range(len(frame)):
                colval = frame.loc[i, 'col']
                index_val = frame.loc[i, 'index']
                src_df.loc[index_val, 'col'] = colval
        new = new.drop('index', axis=1)

        return src_df, new, columdet

    def column_creation_1(self,src_df, columdet):
        """
        This function stores col information of Headers identified
        src_df : Metadata information Dataframe
        columdet : column detection
        """
        header_flag = -1
        frame = src_df.loc[src_df['header'] == 1]
        # check to see if 'HEADER' is detected
        if len(frame) > 0:  # HEADER is present in detection from Column model
            header_flag = 1
            frame.sort_values(['x'], inplace=True)
            frame = frame.reset_index(drop=False)
            for ind in frame.index:
                # ---------- updating 26-Mar-2022 -------------
                # hx1, hx2 will be now decided by combining thr information from both Meta & Cell

                # getting Meta Data Values along X
                metax1 = frame.loc[ind, 'x']
                metax2 = frame.loc[ind, 'x'] + frame.loc[ind, 'w']
                # getting Cell detected info along X-axis
                cellx1 = frame.loc[ind, 'c_x']
                cellx2 = frame.loc[ind, 'c_x'] + frame.loc[ind, 'c_w']
                # --------------------------old logic-----------------------
                # hx1 = frame.loc[ind, 'x']
                # hx2 = frame.loc[ind, 'x'] + frame.loc[ind, 'w']
                # -------------------------------------------------
                hx1 = max(metax1, cellx1)           # deliberately set as Max
                hx2 = min(metax2, cellx2)           # deliberately set as Min
                hy1 = frame.loc[ind, 'y']
                hy2 = frame.loc[ind, 'y'] + frame.loc[ind, 'h']

                for indcol in columdet.index:
                    name = columdet.loc[indcol, 'name']
                    if name not in ['header']:
                        colx1 = columdet.loc[indcol, 'x']
                        colx2 = columdet.loc[indcol, 'x2']
                        coly1 = columdet.loc[indcol, 'y']
                        coly2 = columdet.loc[indcol, 'y2']
                        boo, area, per = self.isRectangleOverlap([hx1, max(hy1, coly1), hx2, min(hy2, coly2)], [colx1, coly1, colx2, coly2])
                        if hx1 >= colx1 and hx2 <= colx2 and frame.loc[ind, 'col'] == -1 and boo is True and per > 0.9:
                            # this checks if value is 100 % inside column
                            frame.loc[ind, 'col'] = indcol
                        else:
                            diff_in_pixel = abs(hx1 - colx1)

                            if hx1 < colx1 and frame.loc[ind, 'col'] == -1:
                                # if value is towards Left of column - x1
                                if diff_in_pixel in range(1, 10) and per > 0.95:
                                    frame.loc[ind, 'col'] = indcol
                                else:
                                    # check the pixel difference
                                    frame.loc[ind, 'col'] = -99
                            if ((colx1 <= hx1 < colx2) or (colx1 < hx2 <= colx2)) and per > 0.9 and frame.loc[ind, 'col'] == -1:
                                # updated 22-MAr-2022
                                # scenarios where value is inside col but width - boundary of col is shotter
                                colx1_next = columdet.loc[indcol + 1, 'x']
                                if hx2 < colx1_next:
                                    frame.loc[ind, 'col'] = indcol
                                    columdet.loc[indcol, 'x2'] = hx2
                                else:
                                    # check the pixel difference
                                    frame.loc[ind, 'col'] = -99

        else:  # if ONLY Column was detected and no Header was found
            src_df.sort_values(['x', 'y'], inplace=True)
            src_df = src_df.reset_index(drop=True)

            columdet.sort_values(['x', 'y'], inplace=True)
            zcol_det_df = columdet.reset_index(drop=True)
            # using metadata here
            for i in range(len(src_df)):
                src_x = src_df.loc[i, 'x']
                src_x2 = src_x + src_df.loc[i, 'w']
                ovelap_lst = {}
                for jind in zcol_det_df.index:
                    col_x = zcol_det_df.loc[jind, 'x']
                    col_x2 = zcol_det_df.loc[jind, 'x2']

                    # check complete overlap
                    if src_x >= col_x and src_x2 <= col_x2 and src_df.loc[i, 'col'] == -1:
                        src_df.loc[i, 'col'] = jind
                        break
                    # check partial Overlap
                    else:
                        overlap = self.is_overlap_check_along_rows(ax1=col_x, ax2=col_x2, aw=col_x2 - col_x, bx1=src_x,
                                                              bx2=src_x2, bw=src_x2 - src_x, perc=0.4)
                        if overlap:
                            boo, per = overlap
                            if boo is True:
                                ovelap_lst.update({jind: per})
                if len(ovelap_lst) > 0:
                    keys = list(ovelap_lst.keys())  # in python 3, you'll need `list(i.keys())`
                    values = list(ovelap_lst.values())
                    max_overlap_was_with_col = keys[values.index(max(values))]
                    src_df.loc[i, 'col'] = max_overlap_was_with_col
            frame = src_df.copy()
        return frame, header_flag

    def column_creation_2(self,frame):
        # this code creates the ctr for unidentified words in header
        ctr = 10
        for i in range(len(frame)):
            val = frame.loc[i, 'col']
            nextval = None
            temp = []
            if val < 0:
                temp.append(i)
                flag = -1
                for j in range(i + 1, len(frame)):
                    jval = frame.loc[j, 'col']
                    if flag < 0 and jval < 0:
                        temp.append(j)
                    else:
                        flag = 1
                        nextval = frame.loc[j, 'col']
                        break
            for k in temp:
                frame.loc[k, 'col'] = ctr
            ctr = ctr + 1
        return frame

    def col_creation_3(self,frame):
        dic = {}
        unique_col = list(frame['col'].unique())
        for e in unique_col:
            temp_df = frame[frame['col'] == e]
            dic.update({e: [min(temp_df['x']), max(temp_df['x'] + temp_df['w'])]})
        new = pd.DataFrame.from_dict(dic)
        new = new.transpose()
        new = new.sort_values(by=[0])
        new = new.reset_index(drop=False)
        return new

    def col_creation_4(self,col_length, col_det_df):
        '''
        this function takes as input
        1. col_length_frame     : this is creteated using min of metadata cordinates and max of metadata
        2. colDetection_frame   : this is the output of Column model
        here we use the two information to adjust the satrt and end values of columns
        '''
        col_length = col_length.sort_values(by=[0])
        col_length = col_length.reset_index(drop=True)

        col_det_df = col_det_df.sort_values(by=['x'])
        col_det_df = col_det_df.reset_index(drop=False)

        # list of all index in Model Detection Frame
        check_list = list(col_det_df['index'])

        for len_ind in col_length.index:
            index_val = col_length.loc[len_ind, 'index']
            len_xmin = col_length.loc[len_ind, 0]
            len_xmax = col_length.loc[len_ind, 1]

            # getting values from col detection model
            if index_val in list(col_det_df['index']):
                det_xmin = col_det_df.loc[index_val, 'x']
                det_xmax = col_det_df.loc[index_val, 'x2']

                xmin = min(len_xmin, det_xmin)
                xmax = max(len_xmax, det_xmax)

                #
                col_length.loc[len_ind, 0] = xmin
                col_length.loc[len_ind, 1] = xmax
                col_det_df.loc[index_val, 'x'] = xmin
                col_det_df.loc[index_val, 'x2'] = xmax
                col_det_df.loc[index_val, 'w'] = xmax - xmin

                check_list.remove(index_val)

            else:
                xmin = len_xmin
                xmax = len_xmax
                col_det_df = col_det_df.append(
                    {'index': index_val, 'x': xmin, 'y': 0, 'w': xmax - xmin, 'h': 0, 'x2': xmax, 'y2': 0,
                     'name': 'column', 'conf': 1}, ignore_index=True)

        if len(check_list) > 0:
            for inde in col_det_df.index:
                index_val = col_det_df.loc[inde, 'index']
                if index_val in check_list:
                    name = col_det_df.loc[inde, 'name']
                    if name == 'header':
                        check_list.remove(index_val)
                    else:
                        xmin = col_det_df.loc[inde, 'x']
                        xmax = col_det_df.loc[inde, 'x2']
                        col_length = col_length.append({'index': index_val, 0: xmin, 1: xmax}, ignore_index=True)

        col_det_df = col_det_df.sort_values(by=['x'])
        col_det_df = col_det_df.reset_index(drop=True)

        col_length = col_length.sort_values(by=[0])
        col_length = col_length.reset_index(drop=True)

        return col_length, col_det_df

    def col_creation_5(self, frame, new):
        lst = []

        for ind in new.index:
            old_val = new.loc[ind, 'index']
            new_val = ind
            for k in range(len(frame)):
                if frame.loc[k, 'col'] == old_val:
                    lst.append(new_val)
        frame['col'] = lst
        return frame

    def column_creation(self,src_df, coldetection):

        # STEP 1. Source Dataframe and detetion of Column Model is used
        # to create a frame where Column 'col' is populated based on
        # a. Header is ther
        # b. No Header was detected
        frame, header_flag = self.column_creation_1(src_df, coldetection)

        # STEP 2. if any column was not identified that is filled
        frame = self.column_creation_2(frame)

        # STEP 3: A Dataframe is created that has width of each column
        new = self.col_creation_3(frame)

        # STEP 4 . Information of Column detection using Model and new_frame in step 3
        # is used to populate proper width of each col
        new, coldetection = self.col_creation_4(new, coldetection)

        # STEP 5 column value is corrected in header frame
        frame = self.col_creation_5(frame, new)

        if header_flag < 0:
            src_df = frame.copy()
        if header_flag > 0:
            for i in range(len(frame)):
                colval = frame.loc[i, 'col']
                index_val = frame.loc[i, 'index']
                src_df.loc[index_val, 'col'] = colval
        new = new.drop('index', axis=1)
        return src_df, new, coldetection

    def column_corr_final(self, src_df, col_range_frame):
        '''
        This function is used to updtate column values
        using detections from Column Model
        src_df              :   Dataframe
        col_range_frame     :      prepared using column detection model
        '''
        for ind in src_df.index:
            colvalue = src_df.loc[ind, 'col']
            if colvalue < 0:
                # using meta data
                metacell_x1 = src_df.loc[ind, 'x']
                metacell_w = src_df.loc[ind, 'w']
                metacell_x2 = metacell_x1 + metacell_w

                for detind in col_range_frame.index:
                    x1 = col_range_frame.loc[detind, 0]
                    x2 = col_range_frame.loc[detind, 1]
                    w = x2 - x1

                    booval = self.is_overlap_check_along_rows(metacell_x1, metacell_x2, metacell_w, x1, x2, w, perc=0.5)
                    if booval:
                        boo, per = booval
                        if boo is True:
                            src_df.loc[ind, 'col'] = detind
                            break
                    else:
                        # Scenarios where metadata did not overlap
                        # using cell value
                        cell_x1 = src_df.loc[ind, 'c_x']
                        cell_w = src_df.loc[ind, 'c_w']
                        cell_x2 = cell_x1 + cell_w

                        for detind in col_range_frame.index:
                            x1 = col_range_frame.loc[detind, 0]
                            x2 = col_range_frame.loc[detind, 1]
                            w = x2 - x1

                            booval = self.is_overlap_check_along_rows(cell_x1, cell_x2, cell_w, x1, x2, w, perc=0.5)
                            if booval:
                                boo, per = booval
                                if boo is True:
                                    src_df.loc[ind, 'col'] = detind
                                    break
        return src_df

    def model_init(self, model_name, det):
        if model_name in ['yolov5']:
            print('\n\t****** Initializing YOLOv5 Architecture for {} ****** '.format(det))
            logger.info('\t****** Initializing YOLOv5 Architecture for {} ****** '.format(det))
            if det.lower() in ['cell', 'cells']:

                # cell model initialization
                # self.cell_model = Yolov5Inference()

                # WEIGHTS and YAML files
                weights_path_cell_detection = str(ROOT) + '/table_cell_config/cell_030122_12_06pm_best.pt'
                yaml_path_cell_detection = str(ROOT) + '/table_cell_config/cells.yaml'
                print('\n****** Looking for Cell Model Files at following path ******')
                print('WEIGHTS file path : ', weights_path_cell_detection)
                print('YAML file path    : ', yaml_path_cell_detection)
                logger.info('\n****** Looking for Cell Model Files at following path ******')
                logger.info('WEIGHTS file path : {} '.format(weights_path_cell_detection))
                logger.info('YAML file path    : {} '.format(yaml_path_cell_detection))
                self.cell_model.model_init(weight_file_path=weights_path_cell_detection,
                                            yaml_file_path=yaml_path_cell_detection)

            elif det.lower() in ['table', 'tables']:

                # table model initialization
                # self.table_model = Yolov5Inference()

                # WEIGHTS and YAML files
                weights_path_table_detection = str(ROOT) + '/table_cell_config/table_030122_1_31am_best.pt'
                yaml_path_table_detection = str(ROOT) + '/table_cell_config/tables.yaml'
                print('\n****** Looking for Table Model Files at following path ****** ')
                print('WEIGHTS file path : ', weights_path_table_detection)
                print('YAML file path    : ', yaml_path_table_detection)
                logger.info('\n****** Looking for Table Model Files at following path ******')
                logger.info('WEIGHTS file path : {} '.format(weights_path_table_detection))
                logger.info('YAML file path    : {} '.format(yaml_path_table_detection))
                self.table_model.model_init(weight_file_path=weights_path_table_detection,
                                            yaml_file_path=yaml_path_table_detection)

        if model_name in ['mmlab', 'mmdet']:
            logger.info('\n\t****** Initializing mmdetection RCNN Architecture for {} ****** '.format(det))
            if det.lower() in ['cell', 'cells']:
                weights_path_cell_detection = str(ROOT) + '/table_cell_config/cell_cascadeRCNN_epoch_50.pth'
                config_cell_detection = str(ROOT) + '/mmdet/configs/cascade_rcnn/cascade_rcnn_r101_fpn_1x_coco_custom_cell.py'

                print('\n****** Looking for Cell Model Files at following path ****** ')
                print('WEIGHTS file path    : ', weights_path_cell_detection)
                print('CONFIG file path     : ', config_cell_detection)

                logger.info('\t****** Looking for Cell Model Files at following path ******')
                logger.info('WEIGHTS file path : {} '.format(weights_path_cell_detection))
                logger.info('YAML file path    : {} '.format(config_cell_detection))

                self.cell_model = init_detector(config_cell_detection, weights_path_cell_detection, device='cpu')

            elif det.lower() in ['table', 'tables']:
                weights_path_table_detection = str(ROOT) + '/table_cell_config/table_cascadeRCNN_epoch_20.pth'
                config_table_detection = str(ROOT) + '/mmdet/configs/cascade_rcnn/cascade_rcnn_r101_fpn_1x_coco_custom.py'

                print('\n****** Looking for Table Model Files at following path ****** ')
                print('WEIGHTS file path    : ', weights_path_table_detection)
                print('CONFIG file path     : ', config_table_detection)

                logger.info('\t****** Looking for Table Model Files at following path ******')
                logger.info('WEIGHTS file path : {} '.format(weights_path_table_detection))
                logger.info('YAML file path    : {} '.format(config_table_detection))

                self.table_model = init_detector(config_table_detection, weights_path_table_detection, device='cpu')

            elif det.lower() in ['column', 'header']:
                weights_path_column_detection = str(ROOT) + '/table_cell_config/epoch_50_column_new.pth'
                config_column_detection = str(ROOT) + '/mmdet/configs/cascade_rcnn/cascade_rcnn_r101_fpn_1x_coco_custom_column.py'

                print('\n******  Looking for Column / Header Model Files at following path ****** ')
                print('WEIGHTS file path    : ', weights_path_column_detection)
                print('CONFIG file path     : ', config_column_detection)

                logger.info('\t****** Looking for Column / Header Model Files at following path ******')
                logger.info('WEIGHTS file path : {} '.format(weights_path_column_detection))
                logger.info('YAML file path    : {} '.format(config_column_detection))

                self.column_model = init_detector(config_column_detection, weights_path_column_detection, device='cpu')

    def restructure_yolov5_predictions(self, table_pred, cell_pred):
        '''
        This code restructures the output of yolo from :
        [x1,y1,x2,y2, conf, clsID] ---> [x1,y1,x2,y2,clsLabel,conf]
        and combines the prediction of cell model and tables model into one final prediction
        list.
        Params:
            table_pred      :  a list of list predictions [[x0, y0. x1, y1, conf, clsId], []]
            cell pred       : a list of list predictions [[x0, y0. x1, y1, conf, clsId], []]
        '''

        # Formatting the output in standard for
        final_pred = []
        for p in table_pred:
            x1, y1, x2, y2, conf, clsid = p
            class_label = self.lookup_category_id_annotation_df.loc[int(clsid), 'name']
            final_pred.append([x1, y1, x2, y2, class_label, round(conf, 2)])

        for p in cell_pred:
            x1, y1, x2, y2, conf, clsid = p
            class_label = self.lookup_category_id_annotation_df.loc[int(clsid), 'name']
            final_pred.append([x1, y1, x2, y2, class_label, round(conf, 2)])

        return final_pred

    def get_detection_yolov5(self, img_ndarray, page_n, object_names_list=['table', 'cell']):
        '''
        Params :
            img_ndarray : input image of type ndaaray
            page_n      : page number in pdf
        Return:
            final_pred  : combined prediction of Table and Cell Model.
        '''

        if ('table' in object_names_list) and self.table_model:
            table_pred = self.table_model.detect_from_ndaaray(img_ndarray=img_ndarray, img_size=640, max_det_=5, plot_thresh_=0.8, save_bbox_img=False)
        if ('cell' in object_names_list) and self.cell_model:
            cell_pred = self.cell_model.detect_from_ndaaray(img_ndarray=img_ndarray, img_size=1024, max_det_=3000, plot_thresh_=0.8, save_bbox_img=False)

        # Restructuring the Detection output of YOLOv5 to standard format used in code
        print('\n\n\t\t For page_n : ', page_n)
        final_pred = self.restructure_yolov5_predictions(table_pred, cell_pred)
        print('Table - Cell prediction is as follows : ', final_pred)

        # Further scope
        # to improve detection
        return final_pred

    def restructure_mmdet_predictions(self, pred, model):
        '''
        pred can be tuple if model is MaskRCNN, and List if model is RCNN
        '''
        if isinstance(pred, tuple):
            bbox_result, segm_result = pred
        else:
            bbox_result, segm_result = pred, None
        bboxes = np.vstack(bbox_result)
        labels = [np.full(bbox.shape[0], i, dtype=np.int32) for i, bbox in enumerate(bbox_result)]
        labels = np.concatenate(labels)

        new_pred = []

        for i, ele in enumerate(bboxes):
            x1 = int(ele[0])
            y1 = int(ele[1])
            x2 = int(ele[2])
            y2 = int(ele[3])
            conf = round(ele[4], 2)
            cls_label = model.CLASSES[labels[i]]

            if cls_label in ['cell', 'special_cell'] and conf > self.cell_det_threshold:
                temp = [x1, y1, x2, y2, cls_label, conf]
                new_pred.append(temp)
            if cls_label in ['gridless_table', 'grided_table', 'semi_grided_table', 'key_value'] and conf > self.table_det_threshold:
                temp = [x1, y1, x2, y2, cls_label, conf]
                new_pred.append(temp)
            else:
                temp = [x1, y1, x2, y2, cls_label, conf]
                new_pred.append(temp)

        return new_pred

    def get_table_v2(self, image_np, page_n, object_names_list=['table', 'cell', 'column']):
        '''
        FOR MMDET detections (MaskRCNN / RCNN)
        This function takes input:
        image_np            : Image in ndarray format.
        page_n              : page number
        object_names_list   : objects we need modle to detect
        Return :
        table_pred          : A List of predictions in table_list_format
        '''

        if ('table' in object_names_list) or ('cell' in object_names_list):
            if ('table' in object_names_list) and self.table_model:
                table_pred = inference_detector(self.table_model, image_np)
                table_pred = self.restructure_mmdet_predictions(table_pred, model=self.table_model)

            if ('cell' in object_names_list) and self.cell_model:
                cell_pred = inference_detector(self.cell_model, image_np)
                cell_pred = self.restructure_mmdet_predictions(cell_pred, model=self.cell_model)

            table_pred.extend(cell_pred)
            return table_pred

        if 'column' in object_names_list:
            column_pred = inference_detector(self.column_model, image_np)
            column_pred = self.restructure_mmdet_predictions(column_pred, model=self.column_model)
            return column_pred

    def metadata_correcction_y_coordnates(self, df):
        z = df.copy()
        z.sort_values(by=['y'], inplace=True)
        z.reset_index(drop=False, inplace=True)

        y_dic = {}

        for i in range(1, len(z)):
            curr_y = z.loc[i, 'y']
            pre_y = z.loc[i - 1, 'y']
            if curr_y == pre_y:
                pass
            else:
                diff = curr_y - pre_y
                if diff > 15:
                    pass
                else:
                    y_dic.update({pre_y: max(curr_y, pre_y)})
                    y_dic.update({curr_y: max(curr_y, pre_y)})

        y = z['y'].replace(y_dic)
        z['y'] = y

        # changing y value too
        for el in y_dic.keys():
            tmp = z.loc[z['y'] == el]
            if len(tmp) > 0:
                h_val = min(tmp['h'])
                for ind in tmp.index:
                    z.loc[ind, 'h'] = h_val

        z.sort_values(by=['index'], inplace=True)
        z.set_index('index', drop=True, inplace=True)
        return z

    def medatdata_y2_values_correction(self, df):

        df_y = df[['y', 'h']]
        df_y.drop_duplicates(inplace=True)
        df_y.sort_values(by=['y'], inplace=True)
        df_y.reset_index(drop=True, inplace=True)
        df_y['y2'] = df_y['y'] + df_y['h']

        df_y_copy = df_y.copy()
        df_y_copy['new_y'] = -1
        df_y_copy['new_y2'] = -1

        for i in range(1, len(df_y)):
            flag = -1
            crr_y = df_y.loc[i, 'y']
            pre_y2 = df_y.loc[i - 1, 'y2']

            if pre_y2 > crr_y:
                new_pre_y2 = min(pre_y2, crr_y) - 3
                new_crr_y = max(pre_y2, crr_y) + 4

                df_y_copy.loc[i, 'new_y'] = new_crr_y
                df_y_copy.loc[i - 1, 'new_y2'] = new_pre_y2

                flag = 1

            # adjusting h value too
            if flag > 0:
                # 1. doing for i
                if df_y_copy.loc[i, 'new_y'] > 0:
                    if df_y_copy.loc[i, 'new_y2'] > 0:
                        df_y_copy.loc[i, 'h'] = df_y_copy.loc[i, 'new_y2'] - df_y_copy.loc[i, 'new_y']
                    else:
                        df_y_copy.loc[i, 'h'] = df_y_copy.loc[i, 'y2'] - df_y_copy.loc[i, 'new_y']
                else:
                    if df_y_copy.loc[i, 'new_y2'] > 0:
                        df_y_copy.loc[i, 'h'] = df_y_copy.loc[i, 'new_y2'] - df_y_copy.loc[i, 'y']

                # 2. doing for i-1
                if df_y_copy.loc[i - 1, 'new_y'] > 0:
                    if df_y_copy.loc[i - 1, 'new_y2'] > 0:
                        df_y_copy.loc[i - 1, 'h'] = df_y_copy.loc[i - 1, 'new_y2'] - df_y_copy.loc[i - 1, 'new_y']
                    else:
                        df_y_copy.loc[i - 1, 'h'] = df_y_copy.loc[i - 1, 'y2'] - df_y_copy.loc[i - 1, 'new_y']
                else:
                    if df_y_copy.loc[i - 1, 'new_y2'] > 0:
                        df_y_copy.loc[i - 1, 'h'] = df_y_copy.loc[i - 1, 'new_y2'] - df_y_copy.loc[i - 1, 'y']

        #  copy of input data frame to mak required changes
        df_copy = df.copy()
        for i in range(len(df_copy)):
            y_val = df_copy.loc[i, 'y']
            for j in range(len(df_y_copy)):
                if y_val == df_y_copy.loc[j, 'y']:
                    df_copy.loc[i, 'h'] = df_y_copy.loc[j, 'h']
                    if df_y_copy.loc[j, 'new_y'] > 0:
                        df_copy.loc[i, 'y'] = df_y_copy.loc[j, 'new_y']
                    break
        return df_copy

    def lookupdetections_y1_y2_same_value_remove(self, lookupdet):
        z1 = lookupdet.copy()
        z1 = z1.drop_duplicates()
        z1.sort_values(by=['x', 'y'], inplace=True)
        z1.reset_index(drop=True, inplace=True)
        z3 = z1.copy()
        for i in range(1, len(z1)):
            curr_y = z1.loc[i, 'y']
            pre_y2 = z1.loc[i - 1, 'y2']
            if curr_y == pre_y2:
                z3.loc[i, 'y'] = curr_y + 2
                z3.loc[i - 1, 'y2'] = pre_y2 - 2

                z3.loc[i, 'h'] = z1.loc[i, 'y2'] - curr_y + 2
                z3.loc[i - 1, 'h'] = (pre_y2 - 2) - z1.loc[i - 1, 'y']

        z3 = z3.drop_duplicates()
        z3.reset_index(drop=True, inplace=True)
        return z3

    def overlap_of_detections_check(self, det):
        '''
        this function returns a detection datafame
        with information of an overlap betwee them
        @parms
            det         : Detecttion Dataframe i.e lookup_detecton_df
        @return
            det_c       : Copy of det df with additional column 'overlap'
                          'overlap' is a str(list of list wih info as :
                          [cell_id, area, per]
        '''
        det_c = det.copy()
        det_c['overlap'] = -1
        for ind in det.index:
            x11 = det.loc[ind, 'x']
            x12 = det.loc[ind, 'x2']
            y11 = det.loc[ind, 'y']
            y12 = det.loc[ind, 'y2']
            id1 = det.loc[ind, 'id']
            cate1 = det.loc[ind, 'supercategory']
            olap = []
            for s_ind in det.index:
                cate2 = det.loc[s_ind, 'supercategory']
                if ind != s_ind and cate1 not in ['table'] and cate2 not in ['table']:

                    x21 = det.loc[s_ind, 'x']
                    x22 = det.loc[s_ind, 'x2']
                    y21 = det.loc[s_ind, 'y']
                    y22 = det.loc[s_ind, 'y2']
                    id2 = det.loc[s_ind, 'id']
                    boo, area, per = self.isRectangleOverlap([x11, y11, x12, y12], [x21, y21, x22, y22])
                    if boo == True:
                        # if thre is an overlap between two
                        olap.append((s_ind, id2, area, per))
            if len(olap) > 0:
                det_c.loc[ind, 'overlap'] = [olap]
        return det_c

    def overlap_det_removal_after_check(self, det, id_ctr=None):

        """
        This function runs post --> overlap_of_detections_check
        if overlap_of_detections_check identifies overlap between 2 cells
        this fumction merges hem with effecive heigh reduce and width of max taken
        """

        remove_index = []
        det_copy = det.copy()

        if not id_ctr:
            id_ctr = len(det_copy)

        tmp_frame = det.loc[det['overlap'] != -1]
        for ind in tmp_frame.index:
            id_ind = tmp_frame.loc[ind, 'id']
            overlap_lst = tmp_frame.loc[ind, 'overlap']
            if id_ind not in remove_index:
                for val in overlap_lst:
                    _, olap_id, _, per = val
                    olap_frma = tmp_frame.loc[tmp_frame['id'] == olap_id]
                    olap_frma.reset_index(drop=False, inplace=True)
                    # remove annotations only if percentage overlap is greater than threshold value
                    if per > 0.8:
                        cr_x = tmp_frame.loc[ind, 'x']
                        cr_y = tmp_frame.loc[ind, 'y']
                        cr_x2 = tmp_frame.loc[ind, 'x2']
                        cr_y2 = tmp_frame.loc[ind, 'y2']
                        cr_h = cr_y2 - cr_y

                        olap_x = olap_frma.loc[0, 'x']
                        olap_y = olap_frma.loc[0, 'y']
                        olap_x2 = olap_frma.loc[0, 'x2']
                        olap_y2 = olap_frma.loc[0, 'y2']
                        olap_h = olap_y2 - olap_y

                        if cr_h > 120 or olap_h > 120:
                            # 120 is an assumed threshold value to know that model detected a mltiline and a cell inside it.
                            # so we wish to keep the multiline and remove the shorter annotation.
                            new_y = min(cr_y, olap_y)
                            new_y2 = max(cr_y2, olap_y2)
                            new_x = min(cr_x, olap_x)
                            new_x2 = max(cr_x2, olap_x2)
                            new_w = new_x2 - new_x
                            new_h = new_y2 - new_y
                        else:
                            new_y = max(cr_y, olap_y)
                            new_y2 = min(cr_y2, olap_y2)
                            new_x = min(cr_x, olap_x)
                            new_x2 = max(cr_x2, olap_x2)
                            new_w = new_x2 - new_x
                            new_h = new_y2 - new_y

                        det_copy = det_copy.append(
                            {'id': id_ctr, 'x': new_x, 'y': new_y, 'w': new_w, 'h': new_h,
                             'x2': new_x2, 'y2': new_y2, 'label': 'cell', 'conf': 2,
                             'category_id': 4, 'name': 'cell',
                             'supercategory': 'sentence', 'overlap': -1},
                            ignore_index=True)
                        id_ctr=id_ctr+1
                        remove_index.append(olap_id)
                        remove_index.append(id_ind)
                    else:
                        det_copy.loc[ind, 'overlap'] = -1

        remove_index = list(set(remove_index))
        for e in remove_index:
            det_copy.drop(det_copy.loc[det_copy['id'] == e].index, inplace=True)

        det_copy.drop(['overlap', 'conf'], axis=1, inplace=True)
        det_copy.set_index(['id'], inplace=True)
        det_copy.drop_duplicates(inplace=True)
        det_copy['id'] = det_copy.index
        det_copy['conf'] = 2
        det_copy['overlap'] = -1
        det_copy.reset_index(drop=True, inplace=True)
        return det_copy

    def identify_overlap_of_det_along_y_axis(self, det, input_df):
        """
        this functon takes lookup_deection as input
        and tries to identify possible multiline that was missed by model
        given atleast 1 multiline  was detected by model

        Combining 2 seperate functions as 1
        """

        # STEP : 1
        '''we identify the y-axis ovelap for eah detection and information is stored in a new colmn'''
        detcopy = det.copy()
        detcopy['y_overlap'] = -1
        detcopy['y_overlap'] = detcopy['y_overlap'].astype(object)

        for i in range(len(detcopy)):
            lst = []
            cur_name = detcopy.loc[i, 'supercategory']
            if cur_name not in ['table']:
                cur_id = detcopy.loc[i, 'id']
                cur_y = detcopy.loc[i, 'y']
                cur_y2 = detcopy.loc[i, 'y2']

                for j in range(len(detcopy)):
                    sub_name = detcopy.loc[j, 'supercategory']
                    if sub_name not in ['table'] and i != j:
                        sub_id = detcopy.loc[j, 'id']
                        sub_y = detcopy.loc[j, 'y']
                        sub_y2 = detcopy.loc[j, 'y2']

                        if (cur_y <= sub_y < cur_y2) or (cur_y < sub_y2 <= cur_y2):
                            # nmber of pixel_overlap calculation
                            if cur_y <= sub_y < cur_y2:
                                num_pixel_overlap = abs(sub_y - cur_y2)
                            if cur_y < sub_y2 <= cur_y2:
                                num_pixel_overlap = abs(sub_y2 - cur_y)

                            if num_pixel_overlap > 10:
                                lst.append([j, sub_id, sub_y, sub_y2, num_pixel_overlap])

                if len(lst) > 0:
                    detcopy.loc[i, 'y_overlap'] = [lst]
                else:
                    pass

        # STEP 2 :
        '''the information captured in above section is use to modify the values of y, y2 and h columns'''

        detcopy2 = detcopy.copy()

        for i in range(len(detcopy2)):
            if detcopy2.loc[i, 'y_overlap'] != -1:
                ids_to_modify = []
                new_y_lst = []
                new_y2_lst = []
                cr_y = detcopy2.loc[i, 'y']
                cr_y2 = detcopy2.loc[i, 'y2']

                new_y_lst.append(cr_y)
                new_y2_lst.append(cr_y2)
                ids_to_modify.append(i)

                y_olap_val = detcopy2.loc[i, 'y_overlap']

                for val in y_olap_val:
                    ind, id_, ol_y, ol_y2, pix_olap = val

                    new_y_lst.append(ol_y)
                    new_y2_lst.append(ol_y2)
                    ids_to_modify.append(ind)

                new_y = min(new_y_lst)
                new_y2 = max(new_y2_lst)
                new_h = new_y2 - new_y

                if new_h > 15:
                    # now changng the vales of y coordinates in detectionn
                    for ides in ids_to_modify:
                        detcopy2.loc[ides, 'y'] = new_y
                        detcopy2.loc[ides, 'y2'] = new_y2
                        detcopy2.loc[ides, 'h'] = new_h

        detcopy2.drop(['y_overlap'], axis=1, inplace=True)
        detcopy3 = self.overlap_of_detections_check(detcopy2)
        detcopy4 = self.overlap_det_removal_after_check(detcopy3, id_ctr=max(detcopy3['id']) + 1)
        detcopy4.dropna(subset=['x'], inplace=True)
        detcopy4.reset_index(drop=True, inplace=True)
        # rerunning cell overlap section again to incorporate new detections
        datadf = self.df_to_table_df_v2(input_df=input_df, det_df=detcopy4)

        return datadf, detcopy4

    def identify_overlap_det_y_axis_multiline(self, src, cell_unique, multi_line_dic):

        """
        this coe helps in idenifying the cell ids wic overlap with cell)ids that have multiline,
        this will help in correcting lines numbes where multiline was detected
        """

        df_copy = src.copy()
        # identify ids where overlpa is present
        olap_dic = {}
        for ids in multi_line_dic.keys():
            lst_ = []
            # cell_values
            tmp_src = df_copy.loc[df_copy['cell_id'] == ids]
            tmp_src.reset_index(drop=True, inplace=True)
            tmp_src_y = tmp_src.loc[0, 'c_y']
            tmp_src_y2 = tmp_src_y + tmp_src.loc[0, 'c_h']

            for cid in cell_unique:
                if cid != ids:

                    tmp_dst = df_copy.loc[df_copy['cell_id'] == cid]
                    tmp_dst.reset_index(drop=True, inplace=True)
                    tmp_dst_y = tmp_dst.loc[0, 'c_y']
                    tmp_dst_y2 = tmp_dst_y + tmp_dst.loc[0, 'c_h']

                    if (tmp_src_y <= tmp_dst_y < tmp_src_y2) or (tmp_src_y < tmp_dst_y2 <= tmp_src_y2):
                        lst_.append(cid)

            if len(lst_) > 0:
                olap_dic.update({ids: lst_})
        return olap_dic

    def pdf_to_page_df_mypypdf_intrim(self, page_num, result_save):
        '''
        This function extrats information for each page
        :param
                page_num        :   Page number of PDF
                result_save     :   Boolean value, used to evaluate if subsequent images generated must be saved or not

        '''

        dataframe_write = None  # Dataframe equivalent for excel
        pixel_dic = dict()  # Local to Page Resolution Info
        page_number = int(page_num) - 1
        page = self.doc.loadPage(page_number)
        pix = page.getPixmap()
        # pix.writePNG(self.result_dir + str(page_num) + '.png')                              # saves  page as PNG file

        x_res = pix.xres  # get x_axis Resolution
        y_res = pix.yres  # get y_axis Resolution
        page_width = pix.width  # get width
        page_height = pix.height  # get height

        # Saving page number into opencv format
        pix = page.getPixmap(matrix=fitz.Matrix(650 / 72, 650 / 72))
        mode = "RGBA" if pix.alpha else "RGB"
        img = Image.frombytes(mode, [pix.width, pix.height], pix.samples)
        size = 2961, 4016
        im_resized = img.resize(size, Image.ANTIALIAS)
        open_cv_image = np.array(im_resized)
        open_cv_image = open_cv_image[:, :, ::-1].copy()  # Convert RGB to BGR

        # converting image to binary
        open_cv_image = cv2.cvtColor(open_cv_image, cv2.COLOR_RGB2GRAY)
        img = np.where(open_cv_image > 200, 255, 0)
        # stacking 1 channel img to 3 channel image &
        open_cv_image = cv2.merge((img, img, img))
        open_cv_image = open_cv_image.astype(np.uint8)
        del img
        image_name = self.result_dir + str(page_num) + '.jpeg'  # file path
        cv2.imwrite(image_name, open_cv_image)  # Saving file

        # 1 page Information update
        pixel_dic.update({'x_res': x_res, 'y_res': y_res, 'width': page_width, 'height': page_height})
        word = page.getTextWords()  # extract text words as a Python list

        labels = ['x0', 'y0', 'x1', 'y1', 'word', 'line_n', 'block_n', 'word_n']
        dataframe = pd.DataFrame.from_records(word, columns=labels)
        dataframe['page_number'] = page_num

        if word:
            w_fact = int(2961) / page_width
            h_fact = int(4016) / page_height

            # ---------------------------------------------------
            if page_width > page_height:
                toRotate = False
                toRotateList = []
                img = Image.frombytes(mode, [pix.width, pix.height], pix.samples)
                size = page_width, page_height  # w-612 , h-792
                img = img.resize(size, Image.ANTIALIAS)
                img = np.array(img)
                img = cv2.cvtColor(img, cv2.COLOR_RGB2GRAY)
                img = np.where(img > 200, 255, 0)
                index_five = dataframe.index[0:5].tolist()
                # cropping for 5 index to check if coordinates need to be rotated or not
                for i in index_five:
                    y0 = int(dataframe.loc[i, 'y0'])
                    y1 = int(dataframe.loc[i, 'y1'])
                    x0 = int(dataframe.loc[i, 'x0'])
                    x1 = int(dataframe.loc[i, 'x1'])
                    crop_img = img[y0:y1, x0:x1]
                    per_num_white_pix = (np.sum(crop_img == 255) / crop_img.size) * 100
                    per_num_black_pix = (np.sum(crop_img == 0) / crop_img.size) * 100
                    if per_num_white_pix > 80:
                        toRotateList.append(True)
                    else:
                        toRotateList.append(False)
                if toRotateList.count(True) > 2:
                    toRotate = True
                if toRotate:
                    print('Rotating the pixel values')
                    # rotating 90 degree clockwise

                    y0_list = dataframe['x0'].tolist()
                    y1_list = dataframe['x1'].tolist()
                    x0_list = dataframe['y1']
                    x1_list = dataframe['y0']

                    x0_list = [792 - i for i in x0_list]
                    x1_list = [792 - i for i in x1_list]

                    # reassigning values
                    dataframe['x0'] = x0_list
                    dataframe['x1'] = x1_list
                    dataframe['y0'] = y0_list
                    dataframe['y1'] = y1_list
            # -------------------------------------------------

            dataframe.rename(columns={'x0': 'x', 'y0': 'y', 'x1': 'w', 'y1': 'h', 'page_number': 'page_number', 'word': 'text'},inplace=True)

            dataframe['w'] = dataframe.apply(lambda row: self.get_width(row['x'], row['w']), axis=1)
            dataframe['h'] = dataframe.apply(lambda row: self.get_heigth(row['y'], row['h']), axis=1)
            dataframe['source'] = dataframe.apply(lambda row: self.get_sourceid(row), axis=1)
            dataframe['x'] = dataframe['x'].apply(lambda x: int(x * w_fact))
            dataframe['w'] = dataframe['w'].apply(lambda x: int(x * w_fact) + 1)
            dataframe['y'] = dataframe['y'].apply(lambda x: int(x * h_fact))
            dataframe['h'] = dataframe['h'].apply(lambda x: int(x * h_fact) + 1)
            # Table
            dataframe['is_in_table'] = 0
            dataframe['table_type'] = ""
            dataframe["table_id"] = -1
            dataframe["t_x"] = -1
            dataframe["t_y"] = -1
            dataframe["t_w"] = -1
            dataframe["t_h"] = -1
            # CELL
            dataframe['is_in_cell'] = 0
            dataframe["cell_id"] = -1
            dataframe['cell_type'] = ""
            dataframe["c_x"] = -1
            dataframe["c_y"] = -1
            dataframe["c_w"] = -1
            dataframe["c_h"] = -1
            # column
            dataframe['header'] = -1
            dataframe['col'] = -1
            # overlap column
            dataframe['cell_overlap'] = -1

            # sorting
            dataframe.sort_values(by=['y', 'x'], inplace=True)
            dataframe.reset_index(drop=True, inplace=True)

            # correcting Meta information in raw frame.
            dataframe = self.metadata_correcction_y_coordnates(dataframe)
            # correcting meta information of overlapping y2 -y coordinates
            dataframe = self.medatdata_y2_values_correction(dataframe)
            self.df_raw = dataframe.copy()

            # ---------------------------------- DEEP LEARNING MODEL FOR DETECTION ---------------------------------
            # DL Model Run   ----- NOTE : table_list has following format : [x0, y0, x1, y1, label, conf]
            # mmlab Model
            table_list = self.get_table_v2(open_cv_image, page_num, object_names_list=['table', 'cell'])  # mmlab model
            # table_list = self.get_detection_maskrcnn(open_cv_image, page_num)                         # MaskRCNN model
            # table_list = self.get_detection_yolov5(open_cv_image, page_num)                            # YOLO v5 model
            # ------------------------------------------------------------------------------------------------------
            # creating a dataframe of all detections made by the model : List ---> Dataframe
            self.lookup_detections_df = self.get_lookup_detection_frame(table_list, self.lookup_category_id_annotation_df)

            # get undetected parts in a IMAGE and SAVE the Masked Image
            # self.get_undetected_parts_img(image_name, table_list, page_num, save_img=result_save)

            # **************************   Column Detection Model  *****************
            self.column_detection = self.column_detection_intrim_pred(open_cv_image, self.lookup_detections_df, page_n=page_num)
            # --- Overlap of columns Remove. # Remember - Column santity is More important than Cell detection Model
            self.column_detection = self.column_det_overlap_remove(self.column_detection)
            self.draw_detetion_save_img(self.column_detection, image_name, page_num, name_of_file='allDetections.jpeg', result_save=result_save, color=(255, 0, 0))

            # ---------------------  Removing all cell annotations that move in multiple columns  -------------------
            self.lookup_detections_df = self.column_remove_annotations_in_multi_columns(cell_det_frame=self.lookup_detections_df, col_det_frame=self.column_detection)

            # ------------------------------------------------------------------------------------------------------
            # fill dataframe with overlap information of table and cells from Model
            dataframe = self.df_to_table_df_v2(input_df=self.df_raw, det_df=self.lookup_detections_df)

            # ------- VERY IMPORTANT :  Check if any table was found by Model or not ------------
            # if not we do not process the file further
            is_table = True if max(dataframe["table_id"]) > -1 else False

            if is_table is True:
                # run if table was detected

                img_buffer = io.BytesIO()
                im_resized.save(img_buffer, dpi=(600, 600), format="jpeg")

                # *********************  UNDETECTED Annotations ***************
                # get undetected annotations of undetected_img and add update the detection dataframe in __init__
                # Dataframe returned below is sanity
                dataframe, self.lookup_detections_df, df_undet_cell_ref = self.get_undetected_parts_bbox(input_df=dataframe, det_df=self.lookup_detections_df, col_det=self.column_detection)
                # dataframe, self.lookup_detections_df = self.height_correction_after_undetected_annotation(df=dataframe, ids_to_work_on=df_undet_cell_ref)
                # self.lookup_detections_df = self.cells_remove_overlap_annotations(col_det_df=self.column_detection, lookup_det=self.lookup_detections_df)
                # self.lookup_detections_df = self.lookupdetections_y1_y2_same_value_remove(lookupdet=self.lookup_detections_df)
                # fill dataframe with overlap information of table and cells AGAIN
                # dataframe = self.df_to_table_df_v2(input_df=dataframe, det_df=self.lookup_detections_df)

                # ------------------ trying to get possible multiline based on detetctions formed -------------
                # Now we again modify the detections based on possible multiline
                # given that model identifies atleast one such mltiline

                dataframe, self.lookup_detections_df = self.identify_overlap_of_det_along_y_axis(det=self.lookup_detections_df, input_df=dataframe)

                image_name_ = self.result_dir + str(page_num) + 'allDetections.jpeg'
                self.draw_detetion_save_img(self.lookup_detections_df, image_name_, page_num, name_of_file='allDetections.jpeg', result_save=result_save, color=(0, 0, 255))

                # --------------------  doing post corrections --------------------------
                # STEP 1 : Identifying values that belong to table and have been successfully detected
                dataframe = dataframe.loc[(dataframe['is_in_cell'] == 1)]
                dataframe = dataframe.loc[(dataframe['is_in_table'] == 1)]

                # STEP 2 : Lines sequence of words are corrected
                dataframe = self.line_num_correction(src_df=dataframe)
                # STEP 3 : Word sequence horizontally are corrected
                dataframe = self.word_num_corr(src_df=dataframe)

                # step 4 : Column Model ran Above
                # b. Idetify headers in the dataset
                dataframe = self.identify_headers(dataframe, self.column_detection)
                # c. Column min_max_creation --> new_col_lengths
                dataframe, new_col_lengths, self.column_detection_new = self.column_creation(dataframe, self.column_detection)
                dataframe = self.column_corr_final(dataframe, new_col_lengths)
                dataframe['block_n'] = dataframe['col']

                # deleting all texts where block value is still -1
                # Get indexes where name column doesn't have value
                indexNames = dataframe[(dataframe['block_n'] == -1)].index
                dataframe.drop(indexNames, inplace=True)

                # STEP 4.a : Handling Multilines detected by model
                orig_dataframe, dataframe = self.identify_cell_ids_with_multiline(src=dataframe)
                # STEP 4.b : Handling Multiline part2
                # since multiline is handled now using 'identify_overlap_of_det_along_y_axis' func above
                # we will not run the below code to reidentify missed multilines
                # dataframe = self.multiline_correction_final(dataframe)

                # -------------------- RULE based multiline handle and EXCEL file write ----------------------
                # dataframe_write is dataframe which is near perfect excel format
                # STEP 5.a : Creating Frame like excel
                dataframe_write = self.formatting_table_to_excel_like(src_frame=dataframe, page_no=page_num)
                # STEP 5.b : Handling Multiline with Rules
                dataframe_write = self.handle_multiline_with_rule(df=dataframe_write)
                # STEP 5.C : Writing this updated frame to excel again
                self.write_excel_after_multiline_handling(src_frame=dataframe_write,
                                                          save_path=self.result_dir + str(self.source) + '_.xlsx',
                                                          read_path=self.result_dir + str(self.source) + '_.xlsx',
                                                          page_no=page_num, save_excel=result_save)

                # shutil.copy(src=self.result_dir + 'pdf_data_excel.xlsx', dst=self.excel_dir_ + str(self.source) + '_.xlsx')
                dataframe_write['page_no'] = page_num
                if self.db_status is None:
                    self.db_status = 'ML_ready'

                return pixel_dic, dataframe, dataframe_write, self.df_raw, self.db_status
            else:
                # logger.error('*** TABLE not detected for Page number : {} *** '.format(page_num))
                if self.db_status is None:
                    self.db_status = 'ML_handled'
                return pixel_dic, dataframe, pd.DataFrame(columns=[0,1,2]), self.df_raw, self.db_status
        else:
            # logger.error('*** TABLE not detected for Page number : {} *** '.format(page_num))
            if self.db_status is None:
                self.db_status = 'ML_holdoff'
            return pixel_dic, dataframe, pd.DataFrame(columns=[0, 1, 2]), dataframe, self.db_status


    def pdf_to_page_df_2(self, pdfpath, UID, page_list, result_save=False, save_result_dir='result'):

        global logger

        self.source = UID

        # creating directory to save Results
        self.result_dir = self.excel_dir + '/' + save_result_dir + '/'
        self.excel_dir_ = self.excel_dir + '/' + save_result_dir + '/'
        os.makedirs(self.result_dir, mode=0o777, exist_ok=True)
        os.makedirs(self.excel_dir_, mode=0o777, exist_ok=True)
        print('\n*** NOTE : Results will be stored in location : ', self.result_dir, '\n')
        logger.info('\t*** NOTE : Results will be stored in location : {} ********\n*'.format(self.result_dir))

        # --------------------------------------  Deep Learning Model Initialization -------------------------------
        if self.table_model is None:
            self.model_init(model_name='mmdet', det='table')
        if self.cell_model is None:
            self.model_init(model_name='mmdet', det='cell')
        if self.column_model is None:
            self.model_init(model_name='mmdet', det='column')
        # ----------------------------------------------------------------------------------------------------------

        extra = {'request_id': 'UniqueRunId:' + str(UID)}
        logger.info(extra)
        logger.info(' Deep Learning Model output has a format as follows : [x0, y0, x1, y1, label, conf] ')
        self.doc = fitz.Document(pdfpath)

        st = 'ML_processed'
        statemet = "UPDATE documentspy SET exception='" + st + "' where documentid=" + str(self.source)
        cursor.execute(statemet)
        cursor.commit()

        page_info = {}
        final_df = pd.DataFrame()
        final_df_raw = pd.DataFrame()
        excel_df = pd.DataFrame()

        response_json = {}
        Table_data = []

        for page_no in page_list:
            logger.info('\n\t\t\t\t\t\t--- Processing Page Number : {} ----\n'.format(page_no))
            print('\t\t\t\t\t page_n : ', page_no)
            # calling the function where detection and processing takes place ---------------------------------
            pixel_dic, dataframe, excel_data, raw_df, dbstatus = self.pdf_to_page_df_mypypdf_intrim(page_num=page_no, result_save=result_save)

            page_info.update({str(page_no): pixel_dic})
            final_df = final_df.append(dataframe, ignore_index=True)
            excel_df = excel_df.append(excel_data, ignore_index=True)
            final_df_raw = final_df_raw.append(raw_df, ignore_index=True)

            # Response
            table_data = excel_data.to_dict()
            Table_data.append({'page_no': page_no,  'table_json': table_data})

        if self.db_status is None:
            self.db_status = 'ML_holdoff'
        self.db_statement = "UPDATE documentspy SET status='" + str(self.db_status) + "' where documentid=" + str(self.source)

        logger.info('\t\t~~~~~~~~~ Status code for the PDF changed to : {} ~~~~~~~~'.format(self.db_status))
        logger.info('\t~~~ SQL update Query  : {} ~~~~'.format(self.db_statement))
        cursor.execute(self.db_statement)
        cursor.commit()

        response_json.update({'page_data': final_df.to_json(orient='split'),
                              'page_excel_data' : excel_df.to_json(orient='split'),
                              'raw_data': final_df_raw.to_json(orient='split'),
                              'Table_data': Table_data,
                              'status_code': 100,
                              "message": "Success..!"})

        return page_info, final_df, final_df_raw, excel_df, response_json, self.db_status


Obj = PyMuPdf()

class table_cell(APIView):

    def post(self, request):
        resp = JResponseJSON()
        try:
            req_data = json.loads(request.body.decode("utf-8"))
            file_path = req_data["PDF_Path"]
            file_path = file_path.replace("\\", "/")
            sid = req_data["Source_ID"]
            page_list= req_data["Page_List"]

            resul_dir = file_path.split('/')[-1].split('.')[0]

            logger.info('\t **** File Path **** : {}'.format(file_path))
            logger.info('\t **** resul_dir **** : {}'.format(resul_dir))
            logger.info('\t **** page_list **** : {}'.format(page_list))

            _,_,_,_,result, dbsat = Obj.pdf_to_page_df_2(pdfpath=file_path, UID=sid, page_list=page_list, result_save=True, save_result_dir='pdf_to_excel_output/'+resul_dir)
            response = JsonResponse(result)
            return response
        except Exception as e:
            logger.error("+++++++++++++++++++++++++++++++++++++++++++++++++++")
            logger.error(req_data)
            logger.error(str(e), exc_info=True)
            logger.error("+++++++++++++++++++++++++++++++++++++++++++++++++++")
            resp.page_data = {}
            resp.raw_data = {}
            resp.Table_data = {}
            resp.status_code = statuscode.FAILURE
            resp.message = "Exception occurred, see the log file."
            response = JsonResponse(resp.__dict__)
            return response


